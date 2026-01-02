"""
FastAPI application for Eternity School Evaluation System.
Provides high-performance API endpoints for evaluation processing.
"""

import csv
import hashlib
import io
import json
import os
import requests
import secrets
from datetime import date, datetime, timedelta
from enum import Enum
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

# Initialize Sentry SDK before FastAPI app
import sentry_sdk
from fastapi import BackgroundTasks, Depends, FastAPI, File, HTTPException, Query, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, ConfigDict, EmailStr, Field, field_validator
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration
from sqlalchemy import func
from sqlalchemy.orm import Session

# MCP integration is optional
try:
    from sentry_sdk.integrations.mcp import MCPIntegration

    MCP_AVAILABLE = True
except (ImportError, Exception):
    MCP_AVAILABLE = False
    MCPIntegration = None

# Sentry configuration
SENTRY_DSN = os.getenv("SENTRY_DSN")  # No default - must be set via environment variable
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")
IS_PRODUCTION = ENVIRONMENT == "production"


def _get_sample_rate(env_value: Optional[str], default_value: float) -> float:
    try:
        return float(env_value) if env_value is not None else default_value
    except ValueError:
        return default_value


DEFAULT_TRACE_RATE = 1.0 if not IS_PRODUCTION else 0.1
DEFAULT_PROFILE_RATE = 1.0 if not IS_PRODUCTION else 0.0

# Initialize Sentry SDK
integrations = [
    FastApiIntegration(),
    SqlalchemyIntegration(),
]
if MCP_AVAILABLE and MCPIntegration:
    integrations.append(MCPIntegration())

# Only initialize Sentry if DSN is provided
if SENTRY_DSN:
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        environment=ENVIRONMENT,
        integrations=integrations,
        traces_sample_rate=_get_sample_rate(os.getenv("SENTRY_TRACES_SAMPLE_RATE"), DEFAULT_TRACE_RATE),
        profile_session_sample_rate=_get_sample_rate(os.getenv("SENTRY_PROFILES_SAMPLE_RATE"), DEFAULT_PROFILE_RATE),
        profile_lifecycle="trace",
        send_default_pii=os.getenv("SENTRY_SEND_DEFAULT_PII", "false").lower() == "true",  # Default to False for security
        release=os.getenv("APP_VERSION", "2.0.0"),
        before_send=lambda event, hint: None if "/health" in event.get("request", {}).get("url", "") else event,
    )
else:
    print("Sentry DSN not set. Skipping Sentry initialization.")

from ai_models.nomination_suggestions import NominationSuggester
from backend.academic_admin_scoring import AcademicAdminScoring
from backend.admin_dashboard import EternitySchoolAdminDashboard
from backend.ai_models.ai_bias_detector import AIBiasDetector
from backend.ai_models.bias_analytics import BiasAnalytics
from backend.ai_models.eom_category_recommender import EOMCategoryRecommender
from backend.audit_logger import AuditLogger
from backend.bias_detection_360 import Complete360BiasDetection
from backend.bulk_import import BulkImporter
from backend.conditional_anonymity_engine import ConditionalAnonymityEngine
from backend.context_aware_bias_detection import ContextAware360BiasDetection
from backend.database import ActionType as DBActionType
from backend.database import (
    Announcement,
    Assignment,
    AuditLog,
    Cycle,
    Database,
    EOMCategory,
    EOMCycle,
    EOMNominee,
    EOMWinner,
    EOMVoter,
    Evaluation,
    Feedback,
    HybridIdentitySession,
    Notification,
    Objection,
    Person,
    StaffSegment,
    Survey,
    SurveyQuestion,
    SurveyResponse,
    SystemSetting,
    VarianceAlert,
    WeightMatrix,
)
from backend.eom_rotation_manager import EOMRotationManager
from backend.eom_validation import EOMNominationValidator
from backend.hybrid_identity_system import HybridIdentityMode, HybridIdentitySurveySystem
from backend.integration_hub import EternitySchoolIntegrationHub
from backend.optimized_evaluation_calculator import OptimizedEvaluationCalculator
from backend.participation_analytics import ParticipationAnalytics
from backend.smart_notification_system import SmartNotificationSystem
from backend.survey_identity_manager import SurveyIdentityManager
from backend.survey_templates import EternitySchoolSurveyTemplates
from backend.system_setup import EternitySchoolSystemSetup
from backend.weight_matrix_handler import WeightMatrixHandler
from backend.auth import SupabaseAuthError, decode_supabase_jwt, extract_bearer_token, fetch_supabase_user

# Initialize FastAPI app
DOCS_ENABLED = os.getenv("ENABLE_DOCS", "false" if IS_PRODUCTION else "true").lower() in ("1", "true", "yes")
docs_url = "/docs" if DOCS_ENABLED else None
redoc_url = "/redoc" if DOCS_ENABLED else None
openapi_url = "/openapi.json" if DOCS_ENABLED else None

app = FastAPI(
    title="Eternity School Evaluation System API",
    description="FastAPI endpoints for evaluation processing, bias detection, and reporting",
    version="2.0.0",
    docs_url=docs_url,
    redoc_url=redoc_url,
    openapi_url=openapi_url,
)

# Initialize task scheduler
import logging

logger = logging.getLogger(__name__)

from backend.task_scheduler import task_scheduler


@app.on_event("startup")
async def startup_event():
    """Start the task scheduler on application startup"""
    # In production we prefer local JWT verification via `SUPABASE_JWT_SECRET`, but we can safely fall back to
    # Supabase's `/auth/v1/user` verification when the secret is missing/misconfigured.
    if IS_PRODUCTION and REQUIRE_SUPABASE_AUTH and not SUPABASE_JWT_SECRET:
        logger.warning(
            "SUPABASE_JWT_SECRET not set; falling back to Supabase /auth/v1/user token verification. "
            "For best performance, set SUPABASE_JWT_SECRET to your project's JWT secret."
        )
    try:
        task_scheduler.start()
        task_scheduler.schedule_daily_tasks()
        logger.info("Task scheduler initialized and started")
    except Exception as e:
        logger.error(f"Error starting task scheduler: {str(e)}")


@app.on_event("shutdown")
async def shutdown_event():
    """Stop the task scheduler on application shutdown"""
    try:
        task_scheduler.stop()
        logger.info("Task scheduler stopped")
    except Exception as e:
        logger.error(f"Error stopping task scheduler: {str(e)}")


# Authentication configuration
REQUIRE_API_KEY = os.getenv("REQUIRE_API_KEY", "false").lower() in ("1", "true", "yes")
API_KEY = os.getenv("ESE_API_KEY") or os.getenv("API_KEY")
REQUIRE_SUPABASE_AUTH = os.getenv("REQUIRE_SUPABASE_AUTH", "true" if IS_PRODUCTION else "false").lower() in (
    "1",
    "true",
    "yes",
)
SUPABASE_JWT_SECRET = os.getenv("SUPABASE_JWT_SECRET") or os.getenv("JWT_SECRET")
SUPABASE_JWT_AUD = (os.getenv("SUPABASE_JWT_AUD") or "authenticated").strip() or None
PUBLIC_PATHS = {
    "/api/v2/health",
    "/api/v2/health/simple",
    "/api/v2/health/config",
    "/api/v2/auth/password-recovery",
    "/docs",
    "/redoc",
    "/openapi.json",
}

def _require_authenticated_email(request: Request) -> str:
    email = getattr(request.state, "user_email", None)
    if not email:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Authentication required")
    return email


def _require_admin_access(request: Request, db: Session) -> str:
    """Admin access (CEO/P&C or super admin)."""
    from backend.rbac_system import RBACSystem

    user_email = _require_authenticated_email(request)
    rbac = RBACSystem(db)
    role = rbac.get_user_role(user_email)
    if not (rbac.is_super_admin(user_email) or role in ("ceo", "pnc")):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user_email


def _hash_session_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _get_hybrid_session(db: Session, session_token: str) -> Optional[HybridIdentitySession]:
    token_hash = _hash_session_token(session_token)
    return db.query(HybridIdentitySession).filter(HybridIdentitySession.token_hash == token_hash).first()


def _store_hybrid_session(
    db: Session,
    session_token: str,
    user_email: str,
    identity_mode: str,
    survey_id: Optional[int],
    permissions: Optional[Dict[str, Any]] = None,
    consent_granted: Optional[Dict[str, Any]] = None,
    expires_at: Optional[datetime] = None,
) -> None:
    token_hash = _hash_session_token(session_token)
    existing = db.query(HybridIdentitySession).filter(HybridIdentitySession.token_hash == token_hash).first()
    if existing:
        existing.user_email = user_email
        existing.identity_mode = identity_mode
        existing.survey_id = survey_id
        existing.permissions = permissions or {}
        existing.consent_granted = consent_granted or {}
        existing.last_activity = datetime.utcnow()
        existing.expires_at = expires_at
        return

    session_row = HybridIdentitySession(
        token_hash=token_hash,
        user_email=user_email,
        identity_mode=identity_mode,
        survey_id=survey_id,
        permissions=permissions or {},
        consent_granted=consent_granted or {},
        last_activity=datetime.utcnow(),
        expires_at=expires_at,
    )
    db.add(session_row)


@app.middleware("http")
async def api_key_middleware(request: Request, call_next):
    """
    Production security gate:
    - Optional API key (x-api-key) for additional hardening
    - Optional Supabase JWT auth (Authorization: Bearer <token>) for user identity
    """
    # Let CORS preflight pass
    if request.method == "OPTIONS":
        return await call_next(request)

    # Public endpoints bypass auth
    if request.url.path in PUBLIC_PATHS:
        return await call_next(request)

    # Optional API key protection
    if REQUIRE_API_KEY:
        if not API_KEY:
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, content={"detail": "API key not configured"}
            )

        provided_key = request.headers.get("x-api-key")
        if not provided_key or not secrets.compare_digest(provided_key, API_KEY):
            return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"detail": "Unauthorized"})

    # Supabase JWT authentication (recommended for production)
    auth_header = request.headers.get("Authorization")
    token = extract_bearer_token(auth_header)

    auth_ctx = None
    if token:
        # 1) Fast path: verify locally when JWT secret is configured.
        if SUPABASE_JWT_SECRET:
            try:
                auth_ctx = decode_supabase_jwt(token, jwt_secret=SUPABASE_JWT_SECRET, audience=SUPABASE_JWT_AUD)
            except SupabaseAuthError:
                auth_ctx = None

        # 2) Fallback: verify against Supabase directly (handles secret mismatch across environments).
        if not auth_ctx:
            supabase_url = (os.getenv("VITE_SUPABASE_URL") or os.getenv("SUPABASE_URL") or "").strip()
            anon_key = (os.getenv("VITE_SUPABASE_ANON_KEY") or os.getenv("SUPABASE_ANON_KEY") or "").strip()
            try:
                auth_ctx = fetch_supabase_user(token, supabase_url=supabase_url, anon_key=anon_key)
            except SupabaseAuthError as exc:
                # If auth is required, fail hard. If not, treat as unauthenticated.
                if REQUIRE_SUPABASE_AUTH:
                    return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"detail": str(exc)})

    if REQUIRE_SUPABASE_AUTH:
        if not token:
            return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"detail": "Authentication required"})
        if not auth_ctx:
            return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"detail": "Invalid token"})

    if auth_ctx:
        request.state.user_email = auth_ctx.email
        request.state.user_id = auth_ctx.user_id
        request.state.auth_role = auth_ctx.role
        request.state.jwt_claims = auth_ctx.claims

    return await call_next(request)


from backend.middleware.logging import StructuredLoggingMiddleware, setup_structured_logging

# Import security middleware after app creation
from backend.middleware.security import RateLimitMiddleware, SecurityHeadersMiddleware

# Redis rate limiting is optional
try:
    from backend.middleware.rate_limit_redis import RedisRateLimitMiddleware

    REDIS_AVAILABLE = True
except (ImportError, Exception):
    REDIS_AVAILABLE = False
    RedisRateLimitMiddleware = None

# Setup structured logging
setup_structured_logging()

# Structured logging middleware (always enabled)
app.add_middleware(StructuredLoggingMiddleware)

# Rate limiting - Use Redis if available, otherwise in-memory
if IS_PRODUCTION:
    rate_limit = int(os.getenv("RATE_LIMIT_PER_MINUTE", "60"))
    use_redis = os.getenv("USE_REDIS_RATE_LIMIT", "false").lower() == "true"

    if use_redis and REDIS_AVAILABLE and RedisRateLimitMiddleware:
        app.add_middleware(RedisRateLimitMiddleware, requests_per_minute=rate_limit)
    else:
        app.add_middleware(RateLimitMiddleware, requests_per_minute=rate_limit)

# Security headers
app.add_middleware(SecurityHeadersMiddleware)

# CORS middleware - Production-safe configuration
ALLOWED_ORIGINS = [origin.strip() for origin in os.getenv("ALLOWED_ORIGINS", "").split(",") if origin.strip()]
# In development, allow localhost; in production, require explicit origins
if IS_PRODUCTION:
    cors_origins = ALLOWED_ORIGINS
else:
    cors_origins = ALLOWED_ORIGINS or ["http://localhost:5173", "http://localhost:3000"]
allow_credentials = "*" not in cors_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=allow_credentials,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)


# Generic error handler (hide details in production)
@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    if IS_PRODUCTION:
        return JSONResponse(status_code=500, content={"detail": "Internal server error"})
    return JSONResponse(status_code=500, content={"detail": str(exc)})


# Sentry debug route for testing (disabled in production)
@app.get("/sentry-debug")
async def trigger_error():
    """Test route to verify Sentry integration"""
    if IS_PRODUCTION:
        raise HTTPException(status_code=404, detail="Not found")
    division_by_zero = 1 / 0


# Database dependency
def get_db():
    """Dependency to get database session"""
    db = Database()
    session = db.get_session()
    try:
        yield session
    finally:
        session.close()


# ============================================================================
# Pydantic Models
# ============================================================================


class EOMNominationRequest(BaseModel):
    """Request model for EOM nomination submission"""

    nominee_email: EmailStr
    eom_cycle_id: int
    nominated_by: Optional[EmailStr] = None  # always derived from auth in production
    nomination_reason: Optional[str] = Field(None, alias="reason")
    category: Optional[EOMCategory] = None
    check_attendance: bool = True

    model_config = ConfigDict(use_enum_values=True, populate_by_name=True)


class EOMNominationResponse(BaseModel):
    """Response model for EOM nomination"""

    nomination_id: int
    is_valid: bool
    errors: List[str]
    warnings: List[str]
    details: Dict[str, Any]


class EOMVoteRequest(BaseModel):
    """Request model for EOM vote submission"""

    eom_cycle_id: int
    nominee_email: EmailStr
    voter_email: Optional[EmailStr] = None  # ignored; derived from auth in production


class MREEvaluationRequest(BaseModel):
    """Request model for MRE evaluation submission"""

    assignment_id: int
    rating: float = Field(..., ge=1.0, le=5.0, description="Rating from 1.0 to 5.0")
    comments: Optional[str] = None
    domain_scores: Optional[Dict[str, float]] = None
    status: str = "submitted"

    @field_validator("status")
    def validate_status(cls, v):
        if v not in ["draft", "submitted", "reviewed"]:
            raise ValueError("Status must be draft, submitted, or reviewed")
        return v


class MREEvaluationResponse(BaseModel):
    """Response model for MRE evaluation"""

    evaluation_id: int
    assignment_id: int
    rating: float
    weighted_rating: float
    weight_applied: float
    target_email: str
    rater_email: str
    target_group: str
    rater_context: str
    status: str


class BiasReportRequest(BaseModel):
    """Request model for bias detection report"""

    cycle_id: int
    include_target_analysis: bool = False
    target_email: Optional[EmailStr] = None


class BiasReportResponse(BaseModel):
    """Response model for bias detection report"""

    cycle_id: int
    overall_bias_score: float
    bias_level: str
    total_evaluations: int
    total_raters: int
    total_targets: int
    findings_count: int
    findings_by_type: Dict[str, int]
    findings_by_severity: Dict[str, int]
    findings: List[Dict[str, Any]]
    context_coverage: Dict[str, Any]
    statistical_summary: Dict[str, Any]
    recommendations: List[str]
    generated_at: str


class CEOReportRequest(BaseModel):
    """Request model for CEO report export"""

    cycle_id: int
    format: str = Field("csv", pattern="^(csv|json|excel)$")
    include_bias_analysis: bool = True
    include_weighted_scores: bool = True
    segment_filter: Optional[StaffSegment] = None


# ============================================================================
# EOM Nomination Endpoints
# ============================================================================

def _resolve_eom_cycle_id(db: Session, cycle_or_eom_cycle_id: int) -> int:
    """
    Accepts either an EOMCycle.id or a Cycle.id and returns a valid EOMCycle.id.

    The frontend historically passes the current Cycle id; for production robustness we:
    - use the provided id directly if it matches an existing EOMCycle
    - otherwise, treat it as Cycle.id and pick (or create) the latest EOMCycle for that Cycle
    """
    # Direct match (EOMCycle.id)
    existing = db.query(EOMCycle).filter(EOMCycle.id == cycle_or_eom_cycle_id).first()
    if existing:
        return existing.id

    # Treat as Cycle.id
    cycle = db.query(Cycle).filter(Cycle.id == cycle_or_eom_cycle_id).first()
    if not cycle:
        raise HTTPException(status_code=404, detail="EOM cycle not found")

    latest = (
        db.query(EOMCycle)
        .filter(EOMCycle.cycle_id == cycle.id)
        .order_by(EOMCycle.year.desc(), EOMCycle.month.desc())
        .first()
    )
    if latest:
        return latest.id

    # Bootstrap a cycle for the current month/year
    now = datetime.utcnow()
    new_cycle = EOMCycle(cycle_id=cycle.id, month=now.month, year=now.year, status="draft")
    db.add(new_cycle)
    db.commit()
    db.refresh(new_cycle)
    return new_cycle.id


def _require_eom_access(request: Request, db: Session) -> str:
    """EOM nominate/vote is restricted to leadership roles."""
    from backend.rbac_system import RBACSystem

    user_email = _require_authenticated_email(request)
    rbac = RBACSystem(db)
    role = rbac.get_user_role(user_email)
    if not (rbac.is_super_admin(user_email) or role in ("ceo", "pnc", "department_head")):
        raise HTTPException(status_code=403, detail="EOM access required")
    return user_email


@app.post("/api/v2/eom/nominations/suggest-category")
async def suggest_eom_category(
    achievement_text: str = Query(..., description="Achievement description or nomination reason"),
    nominee_role: Optional[str] = Query(None, description="Optional role/title of nominee"),
    db: Session = Depends(get_db),
):
    """
    Suggest the most suitable EOM category based on achievement text.
    Uses AI-powered keyword matching and sentiment analysis.
    """
    try:
        recommender = EOMCategoryRecommender()
        result = recommender.suggest_category(achievement_text, nominee_role)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error suggesting category: {str(e)}")


@app.post("/api/v2/eom/nominations/submit", response_model=EOMNominationResponse)
async def submit_eom_nomination(
    nomination: EOMNominationRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """
    Submit an EOM nomination with comprehensive validation.

    Validates:
    - Rotation rules (one win per term)
    - Attendance records
    - Duplicate nominations
    - Leader nomination limits

    If category is not provided, automatically suggests one based on nomination_reason.
    """
    try:
        nominator_email = _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, nomination.eom_cycle_id)

        # Initialize validator and recommender
        validator = EOMNominationValidator(db)
        audit_logger = AuditLogger(db)
        recommender = EOMCategoryRecommender()

        # Auto-suggest category if not provided
        if not nomination.category and nomination.nomination_reason:
            # Get nominee role for better suggestions
            nominee = db.query(Person).filter(Person.email == nomination.nominee_email).first()
            nominee_role = nominee.role_title if nominee else None

            suggestion = recommender.suggest_category(nomination.nomination_reason, nominee_role)
            if suggestion.get("recommended_category") and suggestion.get("confidence_score", 0) > 0.5:
                # Map string category to enum
                try:
                    category_str = suggestion["recommended_category"].upper()
                    # Handle different naming conventions - map to actual enum values
                    category_map = {
                        "OUTSTANDING_LEADERSHIP": EOMCategory.OUTSTANDING_LEADERSHIP,
                        "TEAM_SPIRIT": EOMCategory.TEAM_SPIRIT,
                        "INNOVATION": EOMCategory.INNOVATION,
                        "RISING_STAR": EOMCategory.RISING_STAR,
                        "SERVICE_EXCELLENCE": EOMCategory.SERVICE_EXCELLENCE,
                    }
                    # Also try direct mapping
                    if category_str in category_map:
                        nomination.category = category_map[category_str]
                    else:
                        nomination.category = EOMCategory[category_str]
                except (KeyError, AttributeError, TypeError):
                    pass  # Keep original category if mapping fails

        # Validate nomination
        validation_result = validator.validate_nomination(
            nominee_email=nomination.nominee_email,
            eom_cycle_id=resolved_eom_cycle_id,
            nominated_by=nominator_email,
            category=nomination.category.value if nomination.category else None,
            check_attendance=nomination.check_attendance,
        )

        # If validation fails, return error response
        if not validation_result.is_valid:
            return EOMNominationResponse(
                nomination_id=0,
                is_valid=False,
                errors=validation_result.errors,
                warnings=validation_result.warnings,
                details=validation_result.details,
            )

        # Create nomination if valid
        eom_nominee = EOMNominee(
            eom_cycle_id=resolved_eom_cycle_id,
            nominee_email=nomination.nominee_email,
            nominated_by=nominator_email,
            nomination_reason=nomination.nomination_reason,
            category=nomination.category,  # EOMCategory enum
            rotation_eligible=True,
            votes_received=0,
        )

        db.add(eom_nominee)
        db.commit()
        db.refresh(eom_nominee)

        # Log audit trail
        background_tasks.add_task(
            audit_logger.log_create,
            "eom_nominee",
            eom_nominee.id,
            nominator_email,
            f"Submitted EOM nomination for {nomination.nominee_email}",
        )

        return EOMNominationResponse(
            nomination_id=eom_nominee.id,
            is_valid=True,
            errors=[],
            warnings=validation_result.warnings,
            details={
                **validation_result.details,
                "nomination_id": eom_nominee.id,
                "created_at": eom_nominee.created_at.isoformat() if hasattr(eom_nominee, "created_at") else None,
            },
        )

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error submitting nomination: {str(e)}")


@app.post("/api/v2/eom/nominations/validate")
async def validate_eom_nomination(nomination: EOMNominationRequest, request: Request, db: Session = Depends(get_db)):
    """
    Validate an EOM nomination without submitting it.
    Uses comprehensive rotation rules validation.
    Useful for pre-submission validation.
    """
    try:
        nominator_email = _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, nomination.eom_cycle_id)
        validator = EOMNominationValidator(db)

        validation_result = validator.validate_nomination(
            nominee_email=nomination.nominee_email,
            eom_cycle_id=resolved_eom_cycle_id,
            nominated_by=nominator_email,
            category=nomination.category.value if nomination.category else None,
            check_attendance=nomination.check_attendance,
        )

        return {
            "is_valid": validation_result.is_valid,
            "errors": validation_result.errors,
            "warnings": validation_result.warnings,
            "details": validation_result.details,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error validating nomination: {str(e)}")


@app.post("/api/v2/eom/nominations/batch-validate")
async def validate_batch_nominations(
    nominations: List[EOMNominationRequest],
    request: Request,
    eom_cycle_id: int = Query(..., description="EOM cycle ID for all nominations"),
    db: Session = Depends(get_db),
):
    """Validate multiple EOM nominations at once"""
    try:
        nominator_email = _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, eom_cycle_id)
        validator = EOMNominationValidator(db)

        nomination_dicts = [
            {
                "nominee_email": n.nominee_email,
                "nominated_by": nominator_email,
                "category": n.category.value if n.category else None,
            }
            for n in nominations
        ]

        results = validator.validate_batch_nominations(nomination_dicts, resolved_eom_cycle_id)

        return {
            "total_nominations": len(nominations),
            "valid_count": sum(1 for r in results.values() if r.is_valid),
            "invalid_count": sum(1 for r in results.values() if not r.is_valid),
            "results": {
                email: {"is_valid": result.is_valid, "errors": result.errors, "warnings": result.warnings}
                for email, result in results.items()
            },
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error validating batch nominations: {str(e)}")


@app.get("/api/v2/eom/nominations/cycle/{cycle_id}")
async def list_eom_nominations_for_cycle(cycle_id: int, request: Request, db: Session = Depends(get_db)):
    """List EOM nominations for the current EOM cycle (cycle_id may be Cycle.id or EOMCycle.id)."""
    try:
        _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, cycle_id)

        nominations = (
            db.query(EOMNominee)
            .filter(EOMNominee.eom_cycle_id == resolved_eom_cycle_id)
            .order_by(EOMNominee.created_at.desc())
            .all()
        )

        return [
            {
                "id": n.id,
                "eom_cycle_id": n.eom_cycle_id,
                "nominee_email": n.nominee_email,
                "nominee_name": n.nominee_person.full_name if n.nominee_person else None,
                "nominated_by": n.nominated_by,
                "nomination_reason": n.nomination_reason,
                "category": n.category.value if hasattr(n.category, "value") else n.category,
                "votes_received": n.votes_received,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            }
            for n in nominations
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching nominations: {str(e)}")


@app.post("/api/v2/eom/vote")
async def submit_eom_vote(vote: EOMVoteRequest, request: Request, db: Session = Depends(get_db)):
    """Submit a single EOM vote (one vote per user per EOM cycle)."""
    try:
        voter_email = _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, vote.eom_cycle_id)

        # Prevent double voting
        existing_vote = (
            db.query(EOMVoter)
            .filter(EOMVoter.eom_cycle_id == resolved_eom_cycle_id, EOMVoter.voter_email == voter_email)
            .first()
        )
        if existing_vote:
            raise HTTPException(status_code=400, detail="You have already voted in this EOM cycle")

        nominee = (
            db.query(EOMNominee)
            .filter(EOMNominee.eom_cycle_id == resolved_eom_cycle_id, EOMNominee.nominee_email == vote.nominee_email)
            .first()
        )
        if not nominee:
            raise HTTPException(status_code=404, detail="Nominee not found for this EOM cycle")

        nominee.votes_received = int(nominee.votes_received or 0) + 1

        vote_record = EOMVoter(eom_cycle_id=resolved_eom_cycle_id, voter_email=voter_email)
        db.add(vote_record)
        db.commit()

        return {"success": True, "message": "Vote submitted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error submitting vote: {str(e)}")


@app.get("/api/v2/eom/winners/{cycle_id}")
async def get_eom_winners(cycle_id: int, request: Request, db: Session = Depends(get_db)):
    """Get EOM winners for a cycle (cycle_id may be Cycle.id or EOMCycle.id)."""
    try:
        _require_authenticated_email(request)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, cycle_id)

        winners = db.query(EOMWinner).filter(EOMWinner.eom_cycle_id == resolved_eom_cycle_id).all()
        return [
            {
                "id": w.id,
                "eom_cycle_id": w.eom_cycle_id,
                "winner_email": w.winner_email,
                "category": w.category,
                "term": w.term,
                "votes_received": w.votes_received,
                "announced_at": w.announced_at.isoformat() if w.announced_at else None,
            }
            for w in winners
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching winners: {str(e)}")


# ============================================================================
# EOM Rotation Rule Management Endpoints
# ============================================================================


class RotationRuleRequest(BaseModel):
    """Request model for creating/updating rotation rules"""

    category: EOMCategory
    cycle_id: int
    cooldown_period: int = 3
    max_wins_per_period: int = 1
    period_type: str = "year"  # 'year', 'quarter', 'month'
    is_active: bool = True


class RotationRuleUpdateRequest(BaseModel):
    """Request model for updating rotation rules"""

    cooldown_period: Optional[int] = None
    max_wins_per_period: Optional[int] = None
    period_type: Optional[str] = None
    is_active: Optional[bool] = None


class EligibilityCheckRequest(BaseModel):
    """Request model for checking nominee eligibility"""

    nominee_email: EmailStr
    category: EOMCategory
    eom_cycle_id: int


@app.post("/api/v2/eom/rotation-rules/create")
async def create_rotation_rule(rule: RotationRuleRequest, db: Session = Depends(get_db)):
    """
    Create a new rotation rule for a category.
    """
    try:
        validator = EOMNominationValidator(db)

        created_rule = validator.create_rotation_rule(
            category=rule.category,
            cycle_id=rule.cycle_id,
            cooldown_period=rule.cooldown_period,
            max_wins_per_period=rule.max_wins_per_period,
            period_type=rule.period_type,
            is_active=rule.is_active,
        )

        return {
            "rule_id": created_rule.id,
            "category": created_rule.category.value,
            "cooldown_period": created_rule.cooldown_period,
            "max_wins_per_period": created_rule.max_wins_per_period,
            "period_type": created_rule.period_type,
            "is_active": created_rule.is_active,
            "message": "Rotation rule created successfully",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating rotation rule: {str(e)}")


@app.put("/api/v2/eom/rotation-rules/{rule_id}")
async def update_rotation_rule(rule_id: int, rule_update: RotationRuleUpdateRequest, db: Session = Depends(get_db)):
    """
    Update an existing rotation rule.
    """
    try:
        validator = EOMNominationValidator(db)

        updated_rule = validator.update_rotation_rule(
            rule_id=rule_id,
            cooldown_period=rule_update.cooldown_period,
            max_wins_per_period=rule_update.max_wins_per_period,
            period_type=rule_update.period_type,
            is_active=rule_update.is_active,
        )

        return {
            "rule_id": updated_rule.id,
            "category": updated_rule.category.value,
            "cooldown_period": updated_rule.cooldown_period,
            "max_wins_per_period": updated_rule.max_wins_per_period,
            "period_type": updated_rule.period_type,
            "is_active": updated_rule.is_active,
            "message": "Rotation rule updated successfully",
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating rotation rule: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/cycle/{cycle_id}")
async def get_rotation_rules_for_cycle(
    cycle_id: int,
    category: Optional[EOMCategory] = None,
    active_only: bool = Query(True, description="Only return active rules"),
    db: Session = Depends(get_db),
):
    """
    Get all rotation rules for a cycle.
    """
    try:
        validator = EOMNominationValidator(db)

        rules = validator.get_rotation_rules_for_cycle(cycle_id=cycle_id, category=category, active_only=active_only)

        return {
            "cycle_id": cycle_id,
            "total_rules": len(rules),
            "rules": [
                {
                    "id": rule.id,
                    "category": rule.category.value,
                    "cooldown_period": rule.cooldown_period,
                    "max_wins_per_period": rule.max_wins_per_period,
                    "period_type": rule.period_type,
                    "is_active": rule.is_active,
                    "created_at": rule.created_at.isoformat() if hasattr(rule, "created_at") else None,
                }
                for rule in rules
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving rotation rules: {str(e)}")


@app.post("/api/v2/eom/rotation-rules/check-eligibility")
async def check_nominee_eligibility(request: EligibilityCheckRequest, db: Session = Depends(get_db)):
    """
    Check if a nominee is eligible for nomination based on rotation rules.
    """
    try:
        validator = EOMNominationValidator(db)

        eligibility = validator.check_nominee_rotation_eligibility(
            nominee_email=request.nominee_email, category=request.category, eom_cycle_id=request.eom_cycle_id
        )

        return eligibility
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking eligibility: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/eligible-nominees/{eom_cycle_id}")
async def get_eligible_nominees(
    eom_cycle_id: int,
    request: Request,
    category: Optional[EOMCategory] = Query(None, description="Optional category to check"),
    db: Session = Depends(get_db),
):
    """
    Get list of all eligible nominees for a category in an EOM cycle.
    """
    try:
        _require_eom_access(request, db)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, eom_cycle_id)

        # Frontend currently uses this endpoint as a "nominee directory" (no category filter).
        if category is None:
            people = db.query(Person).filter(Person.active == True).order_by(Person.full_name.asc()).all()
            return [
                {
                    "email": p.email,
                    "full_name": p.full_name,
                    "department": p.department,
                    "role_title": p.role_title,
                    "segment": p.segment.value if hasattr(p.segment, "value") else p.segment,
                }
                for p in people
            ]

        manager = EOMRotationManager(db)
        eligible = manager.get_eligible_nominees(eom_cycle_id=resolved_eom_cycle_id, category=category)
        # Return a flat list for frontend ergonomics.
        return eligible
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving eligible nominees: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/ineligible-nominees/{eom_cycle_id}")
async def get_ineligible_nominees(
    eom_cycle_id: int, category: EOMCategory = Query(..., description="Category to check"), db: Session = Depends(get_db)
):
    """
    Get list of ineligible nominees with reasons.
    """
    try:
        manager = EOMRotationManager(db)

        ineligible = manager.get_ineligible_nominees(eom_cycle_id=eom_cycle_id, category=category)

        return {
            "eom_cycle_id": eom_cycle_id,
            "category": category.value,
            "total_ineligible": len(ineligible),
            "nominees": ineligible,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving ineligible nominees: {str(e)}")


@app.post("/api/v2/eom/rotation-rules/setup-defaults/{cycle_id}")
async def setup_default_rotation_rules(
    cycle_id: int, categories: Optional[List[EOMCategory]] = None, db: Session = Depends(get_db)
):
    """
    Set up default rotation rules for a cycle.
    """
    try:
        manager = EOMRotationManager(db)

        rules = manager.setup_default_rules(cycle_id=cycle_id, categories=categories)

        return {
            "cycle_id": cycle_id,
            "rules_created": len(rules),
            "rules": [
                {
                    "id": rule.id,
                    "category": rule.category.value,
                    "cooldown_period": rule.cooldown_period,
                    "max_wins_per_period": rule.max_wins_per_period,
                    "period_type": rule.period_type,
                }
                for rule in rules
            ],
            "message": f"Created {len(rules)} default rotation rules",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting up default rules: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/analytics/{cycle_id}")
async def get_rotation_analytics(
    cycle_id: int,
    category: Optional[EOMCategory] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
):
    """
    Get comprehensive rotation analytics for a cycle.
    """
    try:
        validator = EOMNominationValidator(db)

        analytics = validator.get_rotation_analytics(
            cycle_id=cycle_id, category=category, start_date=start_date, end_date=end_date
        )

        return analytics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving analytics: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/nominee-history/{nominee_email}")
async def get_nominee_rotation_history(
    nominee_email: str, category: Optional[EOMCategory] = None, db: Session = Depends(get_db)
):
    """
    Get complete rotation history for a nominee.
    """
    try:
        validator = EOMNominationValidator(db)

        history = validator.get_nominee_rotation_history(nominee_email=nominee_email, category=category)

        return history
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving nominee history: {str(e)}")


@app.get("/api/v2/eom/rotation-rules/summary/{cycle_id}")
async def get_rotation_summary(cycle_id: int, db: Session = Depends(get_db)):
    """
    Get comprehensive rotation summary for a cycle.
    """
    try:
        manager = EOMRotationManager(db)

        summary = manager.get_rotation_summary(cycle_id=cycle_id)

        return summary
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving rotation summary: {str(e)}")


@app.post("/api/v2/eom/rotation-rules/update-eligibility-flags/{eom_cycle_id}")
async def update_rotation_eligibility_flags(eom_cycle_id: int, db: Session = Depends(get_db)):
    """
    Update rotation_eligible flags for all nominees in an EOM cycle.
    """
    try:
        manager = EOMRotationManager(db)

        result = manager.update_rotation_eligibility_flags(eom_cycle_id=eom_cycle_id)

        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating eligibility flags: {str(e)}")


# ============================================================================
# MRE Evaluation Endpoints
# ============================================================================


@app.post("/api/v2/mre/evaluations/process", response_model=MREEvaluationResponse)
async def process_mre_evaluation(
    evaluation: MREEvaluationRequest, background_tasks: BackgroundTasks, db: Session = Depends(get_db)
):
    """
    Process an MRE evaluation with automatic weight calculation.

    Features:
    - Applies weight matrix based on target group and rater context
    - Calculates weighted rating
    - Stores domain-specific scores if provided
    """
    try:
        # Get assignment to determine weights
        assignment = db.query(Assignment).filter(Assignment.id == evaluation.assignment_id).first()

        if not assignment:
            raise HTTPException(status_code=404, detail="Assignment not found")

        # Get cycle for weight matrix handler
        cycle = db.query(Cycle).filter(Cycle.id == assignment.cycle_id).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Cycle not found")

        # Initialize weight matrix handler
        weight_handler = WeightMatrixHandler(assignment.cycle_id, db)

        # Get weight for this assignment
        weight = weight_handler.get_weight(target_group=assignment.target_group, rater_context=assignment.rater_context)

        # Calculate weighted rating
        weighted_rating = evaluation.rating * weight

        # Create evaluation record
        eval_record = Evaluation(
            assignment_id=evaluation.assignment_id,
            rating=evaluation.rating,
            weighted_rating=weighted_rating,
            comments=evaluation.comments,
            domain_scores=evaluation.domain_scores,
            status=evaluation.status,
        )

        db.add(eval_record)
        db.commit()
        db.refresh(eval_record)

        # Get rater and target emails
        rater_email = assignment.rater_email
        target_email = assignment.target_email

        # Log audit trail
        audit_logger = AuditLogger(db)
        background_tasks.add_task(
            audit_logger.log_submit,
            "evaluation",
            eval_record.id,
            rater_email,
            f"Processed MRE evaluation for {target_email} with weighted rating {weighted_rating:.2f}",
        )

        return MREEvaluationResponse(
            evaluation_id=eval_record.id,
            assignment_id=evaluation.assignment_id,
            rating=evaluation.rating,
            weighted_rating=weighted_rating,
            weight_applied=weight,
            target_email=target_email,
            rater_email=rater_email,
            target_group=assignment.target_group,
            rater_context=assignment.rater_context,
            status=eval_record.status,
        )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing evaluation: {str(e)}")


@app.get("/api/v2/mre/evaluations/{cycle_id}/weighted-scores")
async def get_weighted_scores(cycle_id: int, target_email: Optional[EmailStr] = None, db: Session = Depends(get_db)):
    """
    Get weighted evaluation scores for a cycle.
    Optionally filter by target email.
    """
    try:
        weight_handler = WeightMatrixHandler(cycle_id, db)

        if target_email:
            # Get scores for specific target
            scores = weight_handler.calculate_final_scores()
            target_scores = scores.get(target_email)

            if not target_scores:
                raise HTTPException(status_code=404, detail=f"No scores found for {target_email}")

            return {"target_email": target_email, "scores": target_scores}
        else:
            # Get all scores
            scores = weight_handler.calculate_final_scores()
            return {"cycle_id": cycle_id, "total_targets": len(scores), "scores": scores}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting weighted scores: {str(e)}")


# ============================================================================
# Bias Detection Endpoints
# ============================================================================


@app.post("/api/v2/bias/reports/generate", response_model=BiasReportResponse)
async def generate_bias_report(request: BiasReportRequest, db: Session = Depends(get_db)):
    """
    Generate comprehensive 360-degree bias detection report.

    Includes:
    - Structural completeness checks
    - Role-based bias
    - Temporal bias
    - Distribution bias
    - Similarity bias
    - ML-based pattern detection
    - Inter-rater reliability
    """
    try:
        detector = Complete360BiasDetection(db)

        # Generate complete report
        report = detector.generate_complete_report(request.cycle_id)

        # Export to dictionary format
        export_data = detector.export_report_to_dict(report)

        # If target-specific analysis requested
        if request.include_target_analysis and request.target_email:
            target_summary = detector.get_bias_summary_by_target(request.cycle_id, request.target_email)
            export_data["target_analysis"] = target_summary

        return BiasReportResponse(**export_data)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating bias report: {str(e)}")


@app.get("/api/v2/bias/reports/{cycle_id}")
async def get_bias_report(
    cycle_id: int,
    include_target_analysis: bool = False,
    target_email: Optional[EmailStr] = None,
    db: Session = Depends(get_db),
):
    """Get bias detection report for a cycle (GET endpoint)"""
    request = BiasReportRequest(cycle_id=cycle_id, include_target_analysis=include_target_analysis, target_email=target_email)
    return await generate_bias_report(request, db)


@app.get("/api/v2/bias/reports/{cycle_id}/target/{target_email}")
async def get_target_bias_summary(cycle_id: int, target_email: EmailStr, db: Session = Depends(get_db)):
    """Get bias summary for a specific target"""
    try:
        detector = Complete360BiasDetection(db)
        summary = detector.get_bias_summary_by_target(cycle_id, target_email)
        return summary
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting target bias summary: {str(e)}")


# ============================================================================
# Context-Aware 360-Degree Bias Detection Endpoints
# ============================================================================


@app.get("/api/v2/bias/analytics/{cycle}")
async def get_bias_analytics(cycle: str, db: Session = Depends(get_db)):
    """
    Generate comprehensive bias analytics report including:
    - Rater fairness scores
    - Department bias heat map
    - Trend analysis (-1 to 1 scale)
    - Mitigation suggestions
    """
    try:
        analytics = BiasAnalytics(db)
        report = analytics.generate_bias_report(cycle)
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating bias analytics: {str(e)}")


@app.post("/api/v2/bias/ai/analyze-evaluation/{evaluation_id}")
async def analyze_evaluation_bias(evaluation_id: int, db: Session = Depends(get_db)):
    """
    Analyze a single evaluation for bias using AI-powered algorithms.

    Returns:
    - bias_indicators: Similarity bias score and confidence
    - bias_flags: List of detected bias types
    - mitigation_suggestions: Specific recommendations
    """
    try:
        evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()

        if not evaluation:
            raise HTTPException(status_code=404, detail="Evaluation not found")

        ai_detector = AIBiasDetector(db)
        result = ai_detector.analyze_evaluation(evaluation)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing evaluation: {str(e)}")


@app.post("/api/v2/bias/ai/analyze-evaluation/{evaluation_id}")
async def analyze_evaluation_bias(evaluation_id: int, db: Session = Depends(get_db)):
    """
    Analyze a single evaluation for bias using AI-powered algorithms.

    Returns:
    - bias_indicators: Similarity bias score and confidence
    - bias_flags: List of detected bias types (e.g., "halo_effect_detected")
    - mitigation_suggestions: Specific recommendations
    """
    try:
        evaluation = db.query(Evaluation).filter(Evaluation.id == evaluation_id).first()

        if not evaluation:
            raise HTTPException(status_code=404, detail="Evaluation not found")

        ai_detector = AIBiasDetector(db)
        result = ai_detector.analyze_evaluation(evaluation)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing evaluation: {str(e)}")


@app.post("/api/v2/bias/360/context-aware-report/{cycle_id}")
async def generate_context_aware_360_report(cycle_id: int, db: Session = Depends(get_db)):
    """
    Generate comprehensive context-aware 360-degree bias detection report.

    Analyzes bias across multiple rater contexts with:
    - Context-specific bias detection
    - Cross-context comparisons
    - Context consistency analysis
    - Context-specific patterns
    - Multi-context statistical analysis
    """
    try:
        detector = ContextAware360BiasDetection(db)

        report = detector.generate_context_aware_report(cycle_id=cycle_id)

        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating context-aware report: {str(e)}")


@app.get("/api/v2/bias/360/context-analysis/{cycle_id}/target/{email}")
async def get_target_context_analysis(cycle_id: int, email: str, db: Session = Depends(get_db)):
    """
    Get detailed context-specific analysis for a target.

    Returns:
    - Ratings breakdown by context
    - Consistency across contexts
    - Missing contexts
    - Context-specific statistics
    """
    try:
        detector = ContextAware360BiasDetection(db)

        analysis = detector.get_target_context_analysis(cycle_id=cycle_id, target_email=email)

        return analysis
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving target context analysis: {str(e)}")


@app.get("/api/v2/bias/360/context-comparison/{cycle_id}")
async def get_context_comparison(
    cycle_id: int,
    context1: Optional[str] = Query(None, description="First context to compare"),
    context2: Optional[str] = Query(None, description="Second context to compare"),
    db: Session = Depends(get_db),
):
    """
    Get detailed comparison between two rater contexts.

    If contexts not specified, returns comparison for all context pairs.
    """
    try:
        detector = ContextAware360BiasDetection(db)

        # Generate full report to get cross-context analyses
        report = detector.generate_context_aware_report(cycle_id=cycle_id)

        if context1 and context2:
            # Filter to specific pair
            comparisons = [
                comp
                for comp in report.get("cross_context_analyses", [])
                if (comp["context_pair"][0] == context1 and comp["context_pair"][1] == context2)
                or (comp["context_pair"][0] == context2 and comp["context_pair"][1] == context1)
            ]
            return {"context_pair": [context1, context2], "comparisons": comparisons}
        else:
            # Return all comparisons
            return {
                "all_comparisons": report.get("cross_context_analyses", []),
                "context_coverage": report.get("context_coverage", {}),
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving context comparison: {str(e)}")


@app.get("/api/v2/bias/360/context-coverage/{cycle_id}")
async def get_context_coverage_analysis(cycle_id: int, db: Session = Depends(get_db)):
    """
    Get context coverage analysis for a cycle.

    Returns:
    - Coverage statistics per context
    - Missing required contexts
    - Balance across contexts
    - Recommendations for improvement
    """
    try:
        detector = ContextAware360BiasDetection(db)

        report = detector.generate_context_aware_report(cycle_id=cycle_id)

        return {
            "context_coverage": report.get("context_coverage", {}),
            "statistical_summary": report.get("statistical_summary", {}),
            "coverage_findings": [
                f
                for f in report.get("findings", [])
                if f.get("bias_type") in ["missing_required_contexts", "insufficient_context_coverage", "context_imbalance"]
            ],
            "recommendations": [
                r for r in report.get("recommendations", []) if "context" in r.lower() or "coverage" in r.lower()
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving context coverage: {str(e)}")


# ============================================================================
# Academic vs Admin Weighted Scoring Endpoints
# ============================================================================


@app.get("/api/v2/scoring/academic-admin/{cycle_id}/score/{email}")
async def get_staff_weighted_score(
    cycle_id: int,
    email: str,
    staff_type: Optional[str] = Query(None, description="Override staff type (academic/admin)"),
    db: Session = Depends(get_db),
):
    """
    Get weighted score for a specific staff member (academic or admin).

    Returns detailed breakdown including:
    - Raw and weighted averages
    - Context-specific scores
    - Final weighted score
    """
    try:
        scorer = AcademicAdminScoring(db)

        score = scorer.calculate_weighted_score(cycle_id=cycle_id, target_email=email, staff_type=staff_type)

        return {
            "cycle_id": cycle_id,
            "target_email": email,
            "staff_type": score.staff_type,
            "total_evaluations": score.total_evaluations,
            "raw_average": score.raw_average,
            "weighted_average": score.weighted_average,
            "final_score": score.final_score,
            "context_breakdown": score.context_breakdown,
            "score_components": score.score_components,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating weighted score: {str(e)}")


@app.get("/api/v2/scoring/academic-admin/{cycle_id}/batch")
async def get_batch_weighted_scores(
    cycle_id: int,
    staff_type: Optional[str] = Query(None, description="Filter by staff type (academic/admin)"),
    target_emails: Optional[str] = Query(None, description="Comma-separated list of emails"),
    db: Session = Depends(get_db),
):
    """
    Get weighted scores for multiple staff members.

    Can filter by staff type or specific emails.
    """
    try:
        scorer = AcademicAdminScoring(db)

        email_list = None
        if target_emails:
            email_list = [e.strip() for e in target_emails.split(",")]

        scores = scorer.calculate_batch_scores(cycle_id=cycle_id, staff_type=staff_type, target_emails=email_list)

        return {
            "cycle_id": cycle_id,
            "staff_type": staff_type or "all",
            "total_scores": len(scores),
            "scores": scorer.export_scores_to_dict(scores),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating batch scores: {str(e)}")


@app.get("/api/v2/scoring/academic-admin/{cycle_id}/compare")
async def compare_academic_vs_admin(cycle_id: int, db: Session = Depends(get_db)):
    """
    Compare scoring between academic and admin staff.

    Returns:
    - Statistics for both staff types
    - Differences and analysis
    - Recommendations for fairness
    """
    try:
        scorer = AcademicAdminScoring(db)

        comparison = scorer.compare_academic_vs_admin(cycle_id=cycle_id)

        return {
            "cycle_id": cycle_id,
            "academic_stats": comparison.academic_stats,
            "admin_stats": comparison.admin_stats,
            "differences": comparison.differences,
            "recommendations": comparison.recommendations,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error comparing academic vs admin: {str(e)}")


@app.get("/api/v2/scoring/academic-admin/{cycle_id}/distribution")
async def get_score_distribution(
    cycle_id: int,
    staff_type: Optional[str] = Query(None, description="Filter by staff type (academic/admin)"),
    db: Session = Depends(get_db),
):
    """
    Get score distribution for academic or admin staff.

    Returns distribution statistics and histogram data.
    """
    try:
        scorer = AcademicAdminScoring(db)

        distribution = scorer.get_score_distribution(cycle_id=cycle_id, staff_type=staff_type)

        return distribution
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting score distribution: {str(e)}")


@app.get("/api/v2/scoring/academic-admin/{cycle_id}/validate")
async def validate_evaluations(
    cycle_id: int,
    staff_type: str = Query(..., description="Staff type to validate (academic/admin)"),
    db: Session = Depends(get_db),
):
    """
    Validate that evaluations meet minimum/maximum requirements for a staff type.

    Returns validation results with errors and warnings.
    """
    try:
        scorer = AcademicAdminScoring(db)

        validation = scorer.validate_evaluations(cycle_id=cycle_id, staff_type=staff_type)

        return validation
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error validating evaluations: {str(e)}")


@app.get("/api/v2/scoring/academic-admin/weight-matrices")
async def get_weight_matrices(db: Session = Depends(get_db)):
    """
    Get the weight matrices for academic and admin staff.

    Returns both matrices for reference.
    """
    try:
        scorer = AcademicAdminScoring(db)

        return {
            "academic": scorer.ACADEMIC_WEIGHT_MATRIX,
            "admin": scorer.ADMIN_WEIGHT_MATRIX,
            "min_evaluations": scorer.MIN_EVALUATIONS,
            "max_evaluations": scorer.MAX_EVALUATIONS,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving weight matrices: {str(e)}")


# ============================================================================
# Optimized Evaluation Calculation Endpoints (for 200+ staff)
# ============================================================================


@app.get("/api/v2/scoring/optimized/batch/{cycle_id}")
async def get_optimized_batch_scores(
    cycle_id: int,
    staff_type: Optional[str] = Query(None, description="Filter by staff type (academic/admin)"),
    target_emails: Optional[str] = Query(None, description="Comma-separated list of emails"),
    db: Session = Depends(get_db),
):
    """
    Get optimized batch scores for multiple staff members (optimized for 200+ staff).

    Uses bulk queries and vectorized operations for maximum performance.
    """
    try:
        calculator = OptimizedEvaluationCalculator(db)

        email_list = None
        if target_emails:
            email_list = [e.strip() for e in target_emails.split(",")]

        scores = calculator.calculate_batch_scores_optimized(
            cycle_id=cycle_id, staff_type=staff_type, target_emails=email_list
        )

        return {
            "cycle_id": cycle_id,
            "staff_type": staff_type or "all",
            "total_scores": len(scores),
            "scores": [
                {
                    "target_email": s.target_email,
                    "staff_type": s.staff_type,
                    "total_evaluations": s.total_evaluations,
                    "raw_average": s.raw_average,
                    "weighted_average": s.weighted_average,
                    "final_score": s.final_score,
                    "context_breakdown": s.context_breakdown,
                }
                for s in scores
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating optimized batch scores: {str(e)}")


@app.get("/api/v2/scoring/optimized/statistics/{cycle_id}")
async def get_optimized_statistics(
    cycle_id: int,
    staff_type: Optional[str] = Query(None, description="Filter by staff type (academic/admin)"),
    db: Session = Depends(get_db),
):
    """
    Get aggregate statistics for all scores in a cycle (optimized for large datasets).
    """
    try:
        calculator = OptimizedEvaluationCalculator(db)

        stats = calculator.get_score_statistics(cycle_id=cycle_id, staff_type=staff_type)

        return {"cycle_id": cycle_id, "staff_type": staff_type or "all", "statistics": stats}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting statistics: {str(e)}")


@app.get("/api/v2/scoring/optimized/compare/{cycle_id}")
async def compare_optimized(cycle_id: int, db: Session = Depends(get_db)):
    """
    Compare academic vs admin scoring using optimized bulk processing.
    """
    try:
        calculator = OptimizedEvaluationCalculator(db)

        comparison = calculator.compare_academic_vs_admin_optimized(cycle_id=cycle_id)

        return comparison
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error comparing scores: {str(e)}")


# ============================================================================
# CEO Report Export Endpoints
# ============================================================================


@app.post("/api/v2/reports/ceo/export")
async def export_ceo_report(request: CEOReportRequest, db: Session = Depends(get_db)):
    """
    Export evaluation results for CEO reports.

    Supports multiple formats:
    - CSV: Comma-separated values
    - JSON: Structured JSON data
    - Excel: Excel spreadsheet (requires openpyxl)

    Includes:
    - Evaluation summaries
    - Weighted scores
    - Bias analysis (optional)
    - Segment breakdowns
    """
    try:
        # Get cycle
        cycle = db.query(Cycle).filter(Cycle.id == request.cycle_id).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Cycle not found")

        # Get all evaluations for this cycle
        evaluations = (
            db.query(Evaluation)
            .join(Assignment)
            .filter(Assignment.cycle_id == request.cycle_id, Evaluation.status == "submitted")
            .all()
        )

        if not evaluations:
            raise HTTPException(status_code=404, detail="No evaluations found for this cycle")

        # Build report data
        report_data = []

        for eval in evaluations:
            assignment = db.query(Assignment).filter(Assignment.id == eval.assignment_id).first()

            if not assignment:
                continue

            # Apply segment filter if specified
            if request.segment_filter:
                target_person = db.query(Person).filter(Person.email == assignment.target_email).first()
                if not target_person or target_person.segment != request.segment_filter:
                    continue

            # Get person details
            target_person = db.query(Person).filter(Person.email == assignment.target_email).first()
            rater_person = db.query(Person).filter(Person.email == assignment.rater_email).first()

            row = {
                "cycle_code": cycle.code,
                "target_email": assignment.target_email,
                "target_name": target_person.full_name if target_person else "Unknown",
                "target_role": target_person.role_title if target_person else "Unknown",
                "target_segment": (
                    target_person.segment.value if target_person and hasattr(target_person, "segment") else "Unknown"
                ),
                "target_group": assignment.target_group,
                "rater_email": assignment.rater_email,
                "rater_name": rater_person.full_name if rater_person else "Unknown",
                "rater_role": rater_person.role_title if rater_person else "Unknown",
                "rater_context": assignment.rater_context,
                "rating": eval.rating,
                "weighted_rating": eval.weighted_rating if eval.weighted_rating else eval.rating,
                "weight": assignment.weight,
                "status": eval.status,
                "submitted_at": eval.submitted_at.isoformat() if eval.submitted_at else None,
            }

            if request.include_weighted_scores and eval.domain_scores:
                row["domain_scores"] = json.dumps(eval.domain_scores)

            report_data.append(row)

        # Generate bias analysis if requested
        bias_data = None
        if request.include_bias_analysis:
            try:
                detector = Complete360BiasDetection(db)
                bias_report = detector.generate_complete_report(request.cycle_id)
                bias_data = detector.export_report_to_dict(bias_report)
            except Exception as e:
                # Don't fail the whole export if bias analysis fails
                bias_data = {"error": str(e)}

        # Format response based on requested format
        if request.format == "csv":
            return generate_csv_response(report_data, bias_data, cycle.code)
        elif request.format == "json":
            return {
                "cycle_id": request.cycle_id,
                "cycle_code": cycle.code,
                "export_date": datetime.utcnow().isoformat(),
                "total_evaluations": len(report_data),
                "evaluations": report_data,
                "bias_analysis": bias_data,
            }
        elif request.format == "excel":
            return generate_excel_response(report_data, bias_data, cycle.code)
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use csv, json, or excel")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting CEO report: {str(e)}")


def generate_csv_response(report_data: List[Dict], bias_data: Optional[Dict], cycle_code: str):
    """Generate CSV response"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=report_data[0].keys() if report_data else [])
    writer.writeheader()
    writer.writerows(report_data)

    # Add bias summary if available
    if bias_data:
        output.write("\n\n=== BIAS ANALYSIS SUMMARY ===\n")
        output.write(f"Overall Bias Score: {bias_data.get('overall_bias_score', 'N/A')}\n")
        output.write(f"Bias Level: {bias_data.get('bias_level', 'N/A')}\n")
        output.write(f"Total Findings: {bias_data.get('findings_count', 0)}\n")

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=ceo_report_{cycle_code}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        },
    )


def generate_excel_response(report_data: List[Dict], bias_data: Optional[Dict], cycle_code: str):
    """Generate Excel response (requires openpyxl)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"

        # Write headers
        if report_data:
            headers = list(report_data[0].keys())
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Write data
            for row in report_data:
                ws.append([row.get(h, "") for h in headers])

        # Add bias analysis sheet if available
        if bias_data:
            ws_bias = wb.create_sheet("Bias Analysis")
            ws_bias.append(["Metric", "Value"])
            ws_bias.append(["Overall Bias Score", bias_data.get("overall_bias_score", "N/A")])
            ws_bias.append(["Bias Level", bias_data.get("bias_level", "N/A")])
            ws_bias.append(["Total Findings", bias_data.get("findings_count", 0)])
            ws_bias.append(["Total Evaluations", bias_data.get("total_evaluations", 0)])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=ceo_report_{cycle_code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
            },
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export requires openpyxl. Install with: pip install openpyxl")


@app.get("/api/v2/reports/ceo/{cycle_id}")
async def get_ceo_report(
    cycle_id: int,
    format: str = Query("json", pattern="^(csv|json|excel)$"),
    include_bias_analysis: bool = True,
    include_weighted_scores: bool = True,
    segment_filter: Optional[StaffSegment] = None,
    db: Session = Depends(get_db),
):
    """Get CEO report (GET endpoint)"""
    request = CEOReportRequest(
        cycle_id=cycle_id,
        format=format,
        include_bias_analysis=include_bias_analysis,
        include_weighted_scores=include_weighted_scores,
        segment_filter=segment_filter,
    )
    return await export_ceo_report(request, db)


# ============================================================================
# Cycles Endpoints
# ============================================================================


class CycleCreateRequest(BaseModel):
    code: str
    name: Optional[str] = None
    start_date: Optional[str] = None  # YYYY-MM-DD
    end_date: Optional[str] = None  # YYYY-MM-DD
    status: Optional[str] = "draft"  # draft, active, closed


class CycleUpdateRequest(BaseModel):
    name: Optional[str] = None
    start_date: Optional[str] = None  # YYYY-MM-DD
    end_date: Optional[str] = None  # YYYY-MM-DD
    status: Optional[str] = None  # draft, active, closed


def _parse_optional_iso_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        raise HTTPException(status_code=422, detail="Invalid date format. Expected YYYY-MM-DD")


@app.get("/api/v2/cycles")
async def get_all_cycles(db: Session = Depends(get_db)):
    """Get all cycles"""
    try:
        cycles = db.query(Cycle).order_by(Cycle.created_at.desc()).all()
        return [
            {
                "id": cycle.id,
                "code": cycle.code,
                "name": cycle.name,
                "start_date": cycle.start_date.isoformat() if cycle.start_date else None,
                "end_date": cycle.end_date.isoformat() if cycle.end_date else None,
                "status": cycle.status,
                "created_at": cycle.created_at.isoformat() if cycle.created_at else None,
                "updated_at": cycle.updated_at.isoformat() if cycle.updated_at else None,
            }
            for cycle in cycles
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving cycles: {str(e)}")


@app.post("/api/v2/cycles")
async def create_cycle(request: CycleCreateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Create a new cycle (admin only)."""
    try:
        _require_admin_access(http_request, db)

        code = str(request.code).strip()
        if not code:
            raise HTTPException(status_code=422, detail="code is required")

        status_value = (request.status or "draft").strip().lower()
        if status_value not in {"draft", "active", "closed"}:
            raise HTTPException(status_code=422, detail="status must be one of: draft, active, closed")

        existing = db.query(Cycle).filter(Cycle.code == code).first()
        if existing:
            raise HTTPException(status_code=409, detail="Cycle code already exists")

        cycle = Cycle(
            code=code,
            name=request.name,
            start_date=_parse_optional_iso_date(request.start_date),
            end_date=_parse_optional_iso_date(request.end_date),
            status=status_value,
        )
        db.add(cycle)

        # If activating, ensure only one active cycle.
        if status_value == "active":
            db.query(Cycle).filter(Cycle.status == "active").update({"status": "draft"}, synchronize_session=False)

        db.commit()
        db.refresh(cycle)

        return {
            "message": "Cycle created successfully",
            "data": {
                "id": cycle.id,
                "code": cycle.code,
                "name": cycle.name,
                "start_date": cycle.start_date.isoformat() if cycle.start_date else None,
                "end_date": cycle.end_date.isoformat() if cycle.end_date else None,
                "status": cycle.status,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating cycle: {str(e)}")


@app.put("/api/v2/cycles/{cycle_id}")
async def update_cycle(
    cycle_id: int, request: CycleUpdateRequest, http_request: Request, db: Session = Depends(get_db)
):
    """Update a cycle (admin only)."""
    try:
        _require_admin_access(http_request, db)

        cycle = db.query(Cycle).filter(Cycle.id == cycle_id).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Cycle not found")

        payload = request.model_dump(exclude_unset=True)
        if "name" in payload:
            cycle.name = payload["name"]
        if "start_date" in payload:
            cycle.start_date = _parse_optional_iso_date(payload["start_date"])
        if "end_date" in payload:
            cycle.end_date = _parse_optional_iso_date(payload["end_date"])
        if "status" in payload and payload["status"] is not None:
            status_value = str(payload["status"]).strip().lower()
            if status_value not in {"draft", "active", "closed"}:
                raise HTTPException(status_code=422, detail="status must be one of: draft, active, closed")
            cycle.status = status_value
            if status_value == "active":
                db.query(Cycle).filter(Cycle.id != cycle_id, Cycle.status == "active").update(
                    {"status": "draft"}, synchronize_session=False
                )

        db.commit()
        db.refresh(cycle)

        return {
            "message": "Cycle updated successfully",
            "data": {
                "id": cycle.id,
                "code": cycle.code,
                "name": cycle.name,
                "start_date": cycle.start_date.isoformat() if cycle.start_date else None,
                "end_date": cycle.end_date.isoformat() if cycle.end_date else None,
                "status": cycle.status,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating cycle: {str(e)}")


@app.get("/api/v2/cycles/current")
async def get_current_cycle(db: Session = Depends(get_db)):
    """Get the current active cycle"""
    try:
        # First try to find a cycle with status 'active'
        cycle = db.query(Cycle).filter(Cycle.status == "active").first()

        # If no active cycle, get the most recent cycle
        if not cycle:
            cycle = db.query(Cycle).order_by(Cycle.created_at.desc()).first()

        # Return null if no cycles exist (frontend can handle this)
        if not cycle:
            return None

        return {
            "id": cycle.id,
            "code": cycle.code,
            "name": cycle.name,
            "start_date": cycle.start_date.isoformat() if cycle.start_date else None,
            "end_date": cycle.end_date.isoformat() if cycle.end_date else None,
            "status": cycle.status,
            "created_at": cycle.created_at.isoformat() if cycle.created_at else None,
            "updated_at": cycle.updated_at.isoformat() if cycle.updated_at else None,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving current cycle: {str(e)}")


@app.get("/api/v2/cycles/{cycle_id}")
async def get_cycle_by_id(cycle_id: int, db: Session = Depends(get_db)):
    """Get a specific cycle by ID"""
    try:
        cycle = db.query(Cycle).filter(Cycle.id == cycle_id).first()
        if not cycle:
            raise HTTPException(status_code=404, detail=f"Cycle {cycle_id} not found")

        return {
            "id": cycle.id,
            "code": cycle.code,
            "name": cycle.name,
            "start_date": cycle.start_date.isoformat() if cycle.start_date else None,
            "end_date": cycle.end_date.isoformat() if cycle.end_date else None,
            "status": cycle.status,
            "created_at": cycle.created_at.isoformat() if cycle.created_at else None,
            "updated_at": cycle.updated_at.isoformat() if cycle.updated_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving cycle: {str(e)}")


# ============================================================================
# People / Staff Endpoints
# ============================================================================


class PersonCreateRequest(BaseModel):
    email: EmailStr
    full_name: str
    role_title: Optional[str] = None
    department: Optional[str] = None
    segment: Optional[str] = "whole_school"  # national, international, whole_school
    hire_date: Optional[str] = None  # YYYY-MM-DD
    active: Optional[bool] = True


class PersonUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    role_title: Optional[str] = None
    department: Optional[str] = None
    segment: Optional[str] = None
    hire_date: Optional[str] = None  # YYYY-MM-DD
    active: Optional[bool] = None


def _parse_staff_segment(value: Optional[str]) -> StaffSegment:
    if not value:
        return StaffSegment.WHOLE_SCHOOL
    try:
        return StaffSegment(str(value).strip().lower())
    except ValueError:
        allowed = ", ".join([s.value for s in StaffSegment])
        raise HTTPException(status_code=422, detail=f"Invalid segment. Allowed: {allowed}")


def _parse_optional_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        raise HTTPException(status_code=422, detail="Invalid date format. Expected YYYY-MM-DD")


@app.get("/api/v2/people")
async def list_people(
    http_request: Request,
    active: Optional[bool] = Query(None),
    segment: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """List people (admin only)."""
    try:
        _require_admin_access(http_request, db)
        query = db.query(Person)
        if active is not None:
            query = query.filter(Person.active == active)
        if segment:
            query = query.filter(Person.segment == _parse_staff_segment(segment))
        people = query.order_by(Person.full_name.asc()).all()
        return [
            {
                "email": p.email,
                "full_name": p.full_name,
                "role_title": p.role_title,
                "department": p.department,
                "segment": p.segment.value if hasattr(p.segment, "value") else p.segment,
                "hire_date": p.hire_date.isoformat() if p.hire_date else None,
                "active": p.active,
                "created_at": p.created_at.isoformat() if p.created_at else None,
                "updated_at": p.updated_at.isoformat() if p.updated_at else None,
            }
            for p in people
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving people: {str(e)}")


@app.get("/api/v2/people/{email}")
async def get_person(email: str, http_request: Request, db: Session = Depends(get_db)):
    """Get a person by email (self or admin)."""
    try:
        current_user_email = _require_authenticated_email(http_request)
        if email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)

        person = db.query(Person).filter(Person.email == email).first()
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")

        return {
            "email": person.email,
            "full_name": person.full_name,
            "role_title": person.role_title,
            "department": person.department,
            "segment": person.segment.value if hasattr(person.segment, "value") else person.segment,
            "hire_date": person.hire_date.isoformat() if person.hire_date else None,
            "active": person.active,
            "created_at": person.created_at.isoformat() if person.created_at else None,
            "updated_at": person.updated_at.isoformat() if person.updated_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving person: {str(e)}")


@app.get("/api/v2/people/segment/{segment}")
async def list_people_by_segment(segment: str, http_request: Request, db: Session = Depends(get_db)):
    """List people by segment (admin only)."""
    try:
        _require_admin_access(http_request, db)
        seg = _parse_staff_segment(segment)
        people = db.query(Person).filter(Person.segment == seg).order_by(Person.full_name.asc()).all()
        return [
            {
                "email": p.email,
                "full_name": p.full_name,
                "role_title": p.role_title,
                "department": p.department,
                "segment": p.segment.value if hasattr(p.segment, "value") else p.segment,
                "hire_date": p.hire_date.isoformat() if p.hire_date else None,
                "active": p.active,
            }
            for p in people
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving people: {str(e)}")


@app.post("/api/v2/people")
async def create_person(request: PersonCreateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Create a new person (admin only)."""
    try:
        _require_admin_access(http_request, db)

        existing = db.query(Person).filter(Person.email == str(request.email)).first()
        if existing:
            raise HTTPException(status_code=409, detail="Person already exists")

        person = Person(
            email=str(request.email),
            full_name=request.full_name,
            role_title=request.role_title,
            department=request.department,
            segment=_parse_staff_segment(request.segment),
            hire_date=_parse_optional_date(request.hire_date),
            active=bool(request.active) if request.active is not None else True,
        )
        db.add(person)
        db.commit()
        db.refresh(person)

        return {
            "message": "Person created successfully",
            "data": {
                "email": person.email,
                "full_name": person.full_name,
                "role_title": person.role_title,
                "department": person.department,
                "segment": person.segment.value if hasattr(person.segment, "value") else person.segment,
                "hire_date": person.hire_date.isoformat() if person.hire_date else None,
                "active": person.active,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating person: {str(e)}")


@app.put("/api/v2/people/{email}")
async def update_person(email: str, request: PersonUpdateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Update an existing person (admin only)."""
    try:
        _require_admin_access(http_request, db)

        person = db.query(Person).filter(Person.email == email).first()
        if not person:
            raise HTTPException(status_code=404, detail="Person not found")

        payload = request.model_dump(exclude_unset=True)
        if "full_name" in payload:
            person.full_name = payload["full_name"]
        if "role_title" in payload:
            person.role_title = payload["role_title"]
        if "department" in payload:
            person.department = payload["department"]
        if "segment" in payload:
            person.segment = _parse_staff_segment(payload["segment"])
        if "hire_date" in payload:
            person.hire_date = _parse_optional_date(payload["hire_date"])
        if "active" in payload:
            person.active = bool(payload["active"])

        db.commit()
        db.refresh(person)

        return {
            "message": "Person updated successfully",
            "data": {
                "email": person.email,
                "full_name": person.full_name,
                "role_title": person.role_title,
                "department": person.department,
                "segment": person.segment.value if hasattr(person.segment, "value") else person.segment,
                "hire_date": person.hire_date.isoformat() if person.hire_date else None,
                "active": person.active,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating person: {str(e)}")


# ============================================================================
# Health Check
# ============================================================================

from backend.monitoring.uptime import uptime_monitor


@app.get("/api/v2/health")
async def health_check(db: Session = Depends(get_db)):
    """Comprehensive health check endpoint for uptime monitoring"""
    return uptime_monitor.get_full_health_check(db)


@app.get("/api/v2/health/simple")
async def health_check_simple():
    """Simple health check (no database dependency)"""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "version": "2.0.0",
    }


@app.get("/api/v2/health/config")
async def health_config():
    """
    Non-secret runtime config flags to avoid guessing in production.
    Safe to expose publicly (booleans only).
    """
    from backend.email_service import EmailService
    from urllib.parse import urlparse

    supabase_url_env = (os.getenv("SUPABASE_URL") or "").strip()
    supabase_url_vite = (os.getenv("VITE_SUPABASE_URL") or "").strip()
    supabase_url = (supabase_url_env or supabase_url_vite or "").strip()

    # Prefer VITE_* fallbacks because they reflect what the frontend is actually configured with.
    anon_key = (os.getenv("SUPABASE_ANON_KEY") or os.getenv("VITE_SUPABASE_ANON_KEY") or "").strip()
    service_role = (os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY") or "").strip()

    email_service = EmailService()
    smtp_password_set = bool(email_service.smtp_password)

    def _host(url: str) -> str:
        if not url:
            return ""
        try:
            parsed = urlparse(url)
            return parsed.netloc or parsed.path
        except Exception:
            return ""

    return {
        "env": os.getenv("ENVIRONMENT", "development"),
        "supabase": {
            "url_set": bool(supabase_url),
            "url_host": _host(supabase_url),
            "url_env_host": _host(supabase_url_env) if supabase_url_env else None,
            "vite_url_host": _host(supabase_url_vite) if supabase_url_vite else None,
            "anon_key_set": bool(anon_key),
            "service_role_key_set": bool(service_role),
            "jwt_secret_set": bool(SUPABASE_JWT_SECRET),
        },
        "smtp": {
            "smtp_server_set": bool(os.getenv("SMTP_SERVER", "").strip()),
            "smtp_user_set": bool(os.getenv("SMTP_USER", "").strip()),
            "smtp_password_set": smtp_password_set,
            "email_enabled_effective": bool(email_service.enabled),
        },
        "password_recovery": {
            "can_generate_link": bool(service_role),
            "can_send_via_smtp": bool(service_role and smtp_password_set),
        },
    }


# ============================================================================
# Auth (Password Recovery)
# ============================================================================


class PasswordRecoveryRequest(BaseModel):
    """Request a password recovery email without leaking whether the account exists."""

    email: EmailStr


class PasswordRecoveryResponse(BaseModel):
    """Generic response for password recovery requests."""

    message: str
    provider: str  # smtp | supabase | none


@app.post("/api/v2/auth/password-recovery", response_model=PasswordRecoveryResponse)
async def password_recovery(payload: PasswordRecoveryRequest, http_request: Request):
    """
    Password recovery flow that prefers SMTP delivery (Resend) when configured.

    Why:
    - Supabase's built-in email delivery can be unreliable in production without custom SMTP.
    - This endpoint uses the Supabase Admin API to generate a recovery link and sends it via SMTP (if configured).
    - If SMTP isn't configured, it falls back to Supabase's `/auth/v1/recover` endpoint.

    Security:
    - Always returns a generic message (no email enumeration).
    - Redirect target is derived from allowed origins (CORS allow-list) to prevent open redirects.
    """

    # Safe redirect target: always use the host we are currently serving (prevents open redirects).
    host = http_request.headers.get("host") or ""
    scheme = "https" if host and "localhost" not in host else "http"
    candidate_origin = f"{scheme}://{host}" if host else "http://localhost:5173"
    redirect_to = f"{candidate_origin.rstrip('/')}/reset-password"

    # Supabase configuration
    supabase_url = (os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL") or "https://ywcfqlyhesnikclesgpr.supabase.co").rstrip(
        "/"
    )
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY")
    anon_key = os.getenv("SUPABASE_ANON_KEY")

    generic_message = "If an account exists for that email, you'll receive a reset link shortly."

    # 1) Preferred: generate recovery link via Supabase Admin API and send via SMTP
    if service_role_key:
        try:
            resp = requests.post(
                f"{supabase_url}/auth/v1/admin/generate_link",
                params={"redirect_to": redirect_to},
                headers={
                    "apikey": service_role_key,
                    "Authorization": f"Bearer {service_role_key}",
                    "Content-Type": "application/json",
                },
                json={"type": "recovery", "email": str(payload.email)},
                timeout=10,
            )

            if resp.ok:
                data = resp.json() if resp.content else {}
                hashed_token = data.get("hashed_token")
                verification_type = data.get("verification_type") or "recovery"

                if hashed_token:
                    # Build a fully-qualified verify URL (properly encoded)
                    verify_url = requests.Request(
                        "GET",
                        f"{supabase_url}/auth/v1/verify",
                        params={"token": hashed_token, "type": verification_type, "redirect_to": redirect_to},
                    ).prepare().url

                    from backend.email_service import EmailService

                    email_service = EmailService()
                    # Only attempt SMTP if SMTP is configured (EmailService will also refuse if disabled)
                    if email_service.smtp_user and email_service.smtp_password:
                        subject = "Reset your password - EVALVision"
                        html_body = f"""
                        <html>
                          <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #0B2B3C;">
                            <div style="max-width: 600px; margin: 0 auto; padding: 24px;">
                              <h2 style="margin: 0 0 12px;">Reset your password</h2>
                              <p style="margin: 0 0 16px;">
                                Click the button below to reset your password. If you did not request this, you can safely ignore this email.
                              </p>
                              <p style="margin: 0 0 24px;">
                                <a href="{verify_url}" style="background: #094773; color: #ffffff; text-decoration: none; padding: 12px 18px; border-radius: 6px; display: inline-block;">
                                  Reset Password
                                </a>
                              </p>
                              <p style="margin: 0; font-size: 12px; color: #4b5563;">
                                If the button doesn't work, copy and paste this link into your browser:<br />
                                <span style="word-break: break-all;">{verify_url}</span>
                              </p>
                            </div>
                          </body>
                        </html>
                        """
                        text_body = f"Reset your password: {verify_url}"
                        sent = email_service.send_email(
                            to_email=str(payload.email),
                            subject=subject,
                            html_body=html_body,
                            text_body=text_body,
                        )
                        if sent:
                            return {"message": generic_message, "provider": "smtp"}
            else:
                # Don't leak details to the client; log minimal info.
                logger.warning("Supabase generate_link failed: status=%s", resp.status_code)
        except Exception as e:
            logger.error("Password recovery (SMTP) attempt failed: %s", str(e))

    # 2) Fallback: Supabase built-in recovery email (requires anon key)
    if anon_key:
        try:
            resp = requests.post(
                f"{supabase_url}/auth/v1/recover",
                headers={"apikey": anon_key, "Content-Type": "application/json"},
                json={"email": str(payload.email), "redirect_to": redirect_to},
                timeout=10,
            )
            if resp.ok:
                return {"message": generic_message, "provider": "supabase"}
            logger.warning("Supabase recover failed: status=%s", resp.status_code)
        except Exception as e:
            logger.error("Password recovery (Supabase) attempt failed: %s", str(e))

    return {"message": generic_message, "provider": "none"}


# ============================================================================
# Objections Endpoints
# ============================================================================


class ObjectionSubmitRequest(BaseModel):
    """Request model for submitting an objection (supports legacy EOM objection payload)"""

    # Legacy EOM objection payload (frontend vote flow)
    eom_nominee_id: Optional[int] = None
    objector_email: Optional[EmailStr] = None  # ignored; derived from auth
    reason: Optional[str] = None

    # Generic objection payload
    submitted_by: Optional[EmailStr] = None  # ignored; derived from auth
    objection_type: Optional[str] = None
    related_entity_type: Optional[str] = None
    related_entity_id: Optional[int] = None
    title: Optional[str] = None
    description: Optional[str] = None


class ObjectionResolveRequest(BaseModel):
    """Request model for resolving an objection"""

    status: str = Field(..., description="Status: resolved, rejected, dismissed")
    resolution_notes: Optional[str] = None
    reviewed_by: Optional[EmailStr] = None


@app.post("/api/v2/objections")
async def submit_objection(request: ObjectionSubmitRequest, http_request: Request, db: Session = Depends(get_db)):
    """Submit an objection to an EOM nomination or evaluation"""
    try:
        current_user_email = _require_authenticated_email(http_request)

        # Normalize legacy payload to generic fields
        if request.eom_nominee_id is not None and request.reason:
            objection_type = "eom_nomination"
            related_entity_type = "eom_nominee"
            related_entity_id = request.eom_nominee_id
            title = "EOM Nomination Objection"
            description = request.reason
        else:
            if not all([request.objection_type, request.related_entity_type, request.related_entity_id, request.title, request.description]):
                raise HTTPException(status_code=400, detail="Missing required objection fields")
            objection_type = request.objection_type
            related_entity_type = request.related_entity_type
            related_entity_id = request.related_entity_id
            title = request.title
            description = request.description

        objection = Objection(
            submitted_by=current_user_email,
            objection_type=objection_type,
            related_entity_type=related_entity_type,
            related_entity_id=related_entity_id,
            title=title,
            description=description,
            status="pending",
        )
        db.add(objection)
        db.commit()
        db.refresh(objection)
        return {"message": "Objection submitted successfully", "id": objection.id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error submitting objection: {str(e)}")


@app.get("/api/v2/objections")
async def get_objections(
    http_request: Request,
    status: Optional[str] = Query(None, description="Filter by status"),
    submitted_by: Optional[str] = Query(None, description="Filter by submitter"),
    db: Session = Depends(get_db),
):
    """Get all objections"""
    try:
        _require_admin_access(http_request, db)
        query = db.query(Objection)
        if status:
            query = query.filter(Objection.status == status)
        if submitted_by:
            query = query.filter(Objection.submitted_by == submitted_by)
        objections = query.order_by(Objection.created_at.desc()).all()

        # Hydrate EOM nominee + submitter names for the admin UI
        def _nominee_name_for(o: Objection) -> Optional[str]:
            if o.related_entity_type != "eom_nominee":
                return None
            nominee = db.query(EOMNominee).filter(EOMNominee.id == o.related_entity_id).first()
            if nominee and nominee.nominee_person:
                return nominee.nominee_person.full_name
            if nominee:
                person = db.query(Person).filter(Person.email == nominee.nominee_email).first()
                return person.full_name if person else nominee.nominee_email
            return None

        def _submitter_name_for(o: Objection) -> Optional[str]:
            person = db.query(Person).filter(Person.email == o.submitted_by).first()
            return person.full_name if person else None

        return [
            {
                "id": o.id,
                # Friendly fields used by the current frontend
                "nominee_name": _nominee_name_for(o),
                "objector_name": _submitter_name_for(o),
                "objector_email": o.submitted_by,
                "reason": o.description,
                "status": o.status,
                "resolution_notes": o.resolution_notes,
                "reviewed_by": o.resolved_by,
                "reviewed_at": o.resolved_at.isoformat() if o.resolved_at else None,
                "created_at": o.created_at.isoformat(),
                # Preserve raw fields for debugging / future UI use
                "submitted_by": o.submitted_by,
                "objection_type": o.objection_type,
                "related_entity_type": o.related_entity_type,
                "related_entity_id": o.related_entity_id,
                "title": o.title,
                "description": o.description,
            }
            for o in objections
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching objections: {str(e)}")


@app.post("/api/v2/objections/{objection_id}/resolve")
async def resolve_objection(
    objection_id: int, request: ObjectionResolveRequest, http_request: Request, db: Session = Depends(get_db)
):
    """Resolve an objection"""
    try:
        current_user_email = _require_admin_access(http_request, db)
        objection = db.query(Objection).filter(Objection.id == objection_id).first()
        if not objection:
            raise HTTPException(status_code=404, detail="Objection not found")

        objection.status = request.status
        objection.resolution_notes = request.resolution_notes
        objection.resolved_by = current_user_email
        objection.resolved_at = datetime.utcnow()

        db.commit()
        return {"message": f"Objection {request.status}", "id": objection_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error resolving objection: {str(e)}")


# ============================================================================
# Announcements Endpoints
# ============================================================================


class AnnouncementCreateRequest(BaseModel):
    """Request model for creating an announcement"""

    title: str = Field(..., min_length=1, max_length=200)
    content: str = Field(..., min_length=1)
    author_email: Optional[EmailStr] = None  # derived from auth in production
    priority: str = Field(default="normal", pattern="^(low|normal|high|urgent)$")
    target_audience: str = Field(default="all", pattern="^(all|ceo|pnc|department_head|staff)$")
    expires_at: Optional[datetime] = None


class AnnouncementUpdateRequest(BaseModel):
    """Request model for updating an announcement"""

    title: Optional[str] = Field(None, min_length=1, max_length=200)
    content: Optional[str] = Field(None, min_length=1)
    priority: Optional[str] = Field(None, pattern="^(low|normal|high|urgent)$")
    target_audience: Optional[str] = Field(None, pattern="^(all|ceo|pnc|department_head|staff)$")
    is_active: Optional[bool] = None
    expires_at: Optional[datetime] = None


@app.post("/api/v2/announcements")
async def create_announcement(request: AnnouncementCreateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Create a new announcement"""
    try:
        from backend.database import Announcement

        current_user_email = _require_admin_access(http_request, db)

        announcement = Announcement(
            title=request.title,
            content=request.content,
            author_email=current_user_email,
            priority=request.priority,
            target_audience=request.target_audience,
            expires_at=request.expires_at,
            is_active=True,
        )
        db.add(announcement)
        db.commit()
        db.refresh(announcement)

        return {
            "message": "Announcement created successfully",
            "id": announcement.id,
            "data": {
                "id": announcement.id,
                "title": announcement.title,
                "content": announcement.content,
                "author_email": announcement.author_email,
                "priority": announcement.priority,
                "target_audience": announcement.target_audience,
                "is_active": announcement.is_active,
                "expires_at": announcement.expires_at.isoformat() if announcement.expires_at else None,
                "created_at": announcement.created_at.isoformat() if announcement.created_at else None,
            },
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating announcement: {str(e)}")


@app.get("/api/v2/announcements")
async def get_announcements(
    http_request: Request,
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    priority: Optional[str] = Query(None, description="Filter by priority"),
    target_audience: Optional[str] = Query(None, description="Filter by target audience"),
    db: Session = Depends(get_db),
):
    """Get all announcements, filtered by active status, priority, and audience"""
    try:
        from datetime import datetime

        from backend.database import Announcement

        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)
        current_role = rbac.get_user_role(current_user_email)
        is_admin = rbac.is_super_admin(current_user_email) or current_role in ("ceo", "pnc")

        query = db.query(Announcement)

        # Filter by active status (default to active only)
        if is_active is None:
            query = query.filter(Announcement.is_active == True)
        elif is_active is not None:
            query = query.filter(Announcement.is_active == is_active)

        # Filter by priority
        if priority:
            query = query.filter(Announcement.priority == priority)

        # Filter by target audience
        if target_audience:
            # Only admins can override the audience filter
            if not is_admin:
                raise HTTPException(status_code=403, detail="Admin access required for audience filtering")
            query = query.filter(Announcement.target_audience == target_audience)
        elif not is_admin:
            # Auto-filter based on current user's inferred role
            query = query.filter(
                (Announcement.target_audience == "all") | (Announcement.target_audience == current_role)
            )

        # Filter out expired announcements
        query = query.filter((Announcement.expires_at.is_(None)) | (Announcement.expires_at > datetime.utcnow()))

        announcements = query.order_by(Announcement.priority.desc(), Announcement.created_at.desc()).all()

        return {
            "data": [
                {
                    "id": a.id,
                    "title": a.title,
                    "content": a.content,
                    "author_email": a.author_email,
                    "author_name": a.author.full_name if a.author else None,
                    "priority": a.priority,
                    "target_audience": a.target_audience,
                    "is_active": a.is_active,
                    "expires_at": a.expires_at.isoformat() if a.expires_at else None,
                    "created_at": a.created_at.isoformat() if a.created_at else None,
                    "updated_at": a.updated_at.isoformat() if a.updated_at else None,
                }
                for a in announcements
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching announcements: {str(e)}")


@app.get("/api/v2/announcements/{announcement_id}")
async def get_announcement(announcement_id: int, http_request: Request, db: Session = Depends(get_db)):
    """Get a single announcement by ID"""
    try:
        from datetime import datetime

        from backend.database import Announcement
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)
        current_role = rbac.get_user_role(current_user_email)
        is_admin = rbac.is_super_admin(current_user_email) or current_role in ("ceo", "pnc")

        query = db.query(Announcement).filter(Announcement.id == announcement_id)
        if not is_admin:
            query = query.filter(Announcement.is_active == True)
            query = query.filter(
                (Announcement.target_audience == "all") | (Announcement.target_audience == current_role)
            )
            query = query.filter((Announcement.expires_at.is_(None)) | (Announcement.expires_at > datetime.utcnow()))

        announcement = query.first()
        if not announcement:
            raise HTTPException(status_code=404, detail="Announcement not found")

        return {
            "data": {
                "id": announcement.id,
                "title": announcement.title,
                "content": announcement.content,
                "author_email": announcement.author_email,
                "author_name": announcement.author.full_name if announcement.author else None,
                "priority": announcement.priority,
                "target_audience": announcement.target_audience,
                "is_active": announcement.is_active,
                "expires_at": announcement.expires_at.isoformat() if announcement.expires_at else None,
                "created_at": announcement.created_at.isoformat() if announcement.created_at else None,
                "updated_at": announcement.updated_at.isoformat() if announcement.updated_at else None,
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching announcement: {str(e)}")


@app.put("/api/v2/announcements/{announcement_id}")
async def update_announcement(
    announcement_id: int, request: AnnouncementUpdateRequest, http_request: Request, db: Session = Depends(get_db)
):
    """Update an announcement"""
    try:
        from backend.database import Announcement

        _require_admin_access(http_request, db)

        announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
        if not announcement:
            raise HTTPException(status_code=404, detail="Announcement not found")

        if request.title is not None:
            announcement.title = request.title
        if request.content is not None:
            announcement.content = request.content
        if request.priority is not None:
            announcement.priority = request.priority
        if request.target_audience is not None:
            announcement.target_audience = request.target_audience
        if request.is_active is not None:
            announcement.is_active = request.is_active
        if request.expires_at is not None:
            announcement.expires_at = request.expires_at

        db.commit()
        db.refresh(announcement)

        return {
            "message": "Announcement updated successfully",
            "data": {
                "id": announcement.id,
                "title": announcement.title,
                "content": announcement.content,
                "priority": announcement.priority,
                "target_audience": announcement.target_audience,
                "is_active": announcement.is_active,
                "expires_at": announcement.expires_at.isoformat() if announcement.expires_at else None,
                "updated_at": announcement.updated_at.isoformat() if announcement.updated_at else None,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating announcement: {str(e)}")


@app.delete("/api/v2/announcements/{announcement_id}")
async def delete_announcement(announcement_id: int, http_request: Request, db: Session = Depends(get_db)):
    """Delete an announcement (soft delete by setting is_active=False)"""
    try:
        from backend.database import Announcement

        _require_admin_access(http_request, db)

        announcement = db.query(Announcement).filter(Announcement.id == announcement_id).first()
        if not announcement:
            raise HTTPException(status_code=404, detail="Announcement not found")

        announcement.is_active = False
        db.commit()

        return {"message": "Announcement deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting announcement: {str(e)}")


# ============================================================================
# Notifications Endpoints
# ============================================================================


@app.get("/api/v2/notifications")
async def get_notifications(
    request: Request,
    user_email: Optional[str] = Query(None, description="(Admin only) fetch notifications for a specific user"),
    read: Optional[bool] = Query(None, description="Filter by read status"),
    db: Session = Depends(get_db),
):
    """Get notifications for a user"""
    try:
        current_user_email = _require_authenticated_email(request)
        target_email = user_email or current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)

        query = db.query(Notification).filter(Notification.recipient_email == target_email)
        if read is not None:
            query = query.filter(Notification.read == read)
        notifications = query.order_by(Notification.created_at.desc()).all()
        return [
            {
                "id": n.id,
                "recipient_email": n.recipient_email,
                "notification_type": n.notification_type,
                "title": n.title,
                "message": n.message,
                "read": n.read,
                "read_at": n.read_at.isoformat() if n.read_at else None,
                "action_url": n.action_url,
                "related_entity_type": n.related_entity_type,
                "related_entity_id": n.related_entity_id,
                "priority": n.priority,
                "created_at": n.created_at.isoformat(),
            }
            for n in notifications
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    request: Request,
    user_email: Optional[str] = Query(None, description="(Admin only) mark a notification for a specific user"),
    db: Session = Depends(get_db),
):
    """Mark notification as read"""
    try:
        current_user_email = _require_authenticated_email(request)
        target_email = user_email or current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)

        notification = (
            db.query(Notification)
            .filter(Notification.id == notification_id, Notification.recipient_email == target_email)
            .first()
        )
        if not notification:
            raise HTTPException(status_code=404, detail="Notification not found")

        notification.read = True
        notification.read_at = datetime.utcnow()
        db.commit()
        return {"message": "Notification marked as read", "id": notification_id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/notifications/read-all")
async def mark_all_notifications_read(
    request: Request,
    user_email: Optional[str] = Query(None, description="(Admin only) mark all notifications for a specific user"),
    db: Session = Depends(get_db),
):
    """Mark all notifications as read for a user"""
    try:
        current_user_email = _require_authenticated_email(request)
        target_email = user_email or current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)

        updated = (
            db.query(Notification)
            .filter(Notification.recipient_email == target_email, Notification.read == False)
            .update({"read": True, "read_at": datetime.utcnow()}, synchronize_session=False)
        )
        db.commit()
        return {"message": f"Marked {updated} notifications as read"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/notifications/user-behavior/{user_email}")
async def get_user_behavior_profile(user_email: str, request: Request, db: Session = Depends(get_db)):
    """
    Get user behavior profile for smart reminder timing.
    Returns optimal reminder times based on historical completion patterns.
    """
    try:
        current_user_email = _require_authenticated_email(request)
        if user_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)
        notification_system = SmartNotificationSystem(db)
        profile = notification_system.get_user_behavior_profile(user_email)
        return profile
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting behavior profile: {str(e)}")


@app.post("/api/v2/notifications/smart-reminder/{cycle_id}")
async def send_smart_reminders(
    cycle_id: int,
    request: Request,
    user_email: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Send smart reminders based on user behavior profiles.
    Only sends if current time is optimal for the user.
    """
    try:
        _require_admin_access(request, db)
        from backend.database import Assignment, Evaluation

        notification_system = SmartNotificationSystem(db)

        # Get pending evaluations
        query = (
            db.query(Assignment)
            .outerjoin(Evaluation, Assignment.id == Evaluation.assignment_id)
            .filter(Assignment.cycle_id == cycle_id, Evaluation.id.is_(None))
        )

        if user_email:
            query = query.filter(Assignment.rater_email == user_email)

        pending_assignments = query.all()

        if not pending_assignments:
            return {"success": True, "message": "No pending evaluations", "reminders_sent": 0}

        # Group by rater
        assignments_by_rater = {}
        for assignment in pending_assignments:
            if assignment.rater_email not in assignments_by_rater:
                assignments_by_rater[assignment.rater_email] = []
            assignments_by_rater[assignment.rater_email].append(
                {"assignment_id": assignment.id, "target_email": assignment.target_email}
            )

        results = {"reminders_sent": 0, "reminders_skipped": 0, "details": []}

        for rater_email, assignments in assignments_by_rater.items():
            reminder_result = notification_system.send_smart_reminder(
                user_email=rater_email, pending_evaluations=assignments, cycle_id=cycle_id
            )

            if reminder_result.get("success"):
                results["reminders_sent"] += 1
            else:
                results["reminders_skipped"] += 1

            results["details"].append({"rater_email": rater_email, "result": reminder_result})

        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error sending smart reminders: {str(e)}")


@app.post("/api/v2/notifications/check-overdue/{cycle_id}")
async def check_overdue_evaluations(
    cycle_id: int,
    request: Request,
    escalation_days: int = Query(7, ge=1, le=30),
    db: Session = Depends(get_db),
):
    """
    Check for overdue evaluations and send escalation alerts.

    Args:
        cycle_id: Evaluation cycle ID
        escalation_days: Days after which to escalate (default: 7)
    """
    try:
        _require_admin_access(request, db)
        notification_system = SmartNotificationSystem(db)
        result = notification_system.check_overdue_evaluations(cycle_id=cycle_id, escalation_days=escalation_days)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking overdue evaluations: {str(e)}")


@app.post("/api/v2/notifications/due-soon-reminders/{cycle_id}")
async def send_due_soon_reminders(
    cycle_id: int,
    request: Request,
    days_before: int = Query(3, ge=1, le=14),
    db: Session = Depends(get_db),
):
    """
    Send reminders for evaluations due soon, using smart timing based on user behavior.

    Args:
        cycle_id: Evaluation cycle ID
        days_before: Days before deadline to send reminder (default: 3)
    """
    try:
        _require_admin_access(request, db)
        notification_system = SmartNotificationSystem(db)
        result = notification_system.send_due_soon_reminders(cycle_id=cycle_id, days_before=days_before)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error sending due soon reminders: {str(e)}")


# ============================================================================
# Survey Identity Management Endpoints
# ============================================================================


class IdentityPreferenceRequest(BaseModel):
    """Request model for setting identity preference"""

    user_email: Optional[EmailStr] = None  # derived from auth in production
    preference: str = Field(..., description="Identity mode: 'anonymous', 'identified', or 'conditional'")
    survey_id: Optional[int] = None


class IdentityRevealRequest(BaseModel):
    """Request model for identity reveal"""

    user_email: Optional[EmailStr] = None  # derived from auth in production
    method: str = Field(
        ..., description="Reveal method: 'full', 'partial_role', 'partial_department', 'gradual', 'consent_based'"
    )
    target: Optional[str] = None
    consent: Optional[bool] = False
    conditions: Optional[Dict[str, Any]] = None
    revoke_anonymity: Optional[bool] = False
    survey_id: Optional[int] = None


@app.post("/api/v2/survey/identity/preference")
async def set_identity_preference(request: IdentityPreferenceRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Set user identity preference for surveys.

    Options:
    - anonymous: Complete anonymity, no identity tracking
    - identified: Full identification, standard tracking
    - conditional: Conditional reveal with consent-based options
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        identity_manager = SurveyIdentityManager(db)
        result = identity_manager.set_identity_preference(
            user_id=current_user_email, preference=request.preference, survey_id=request.survey_id
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting identity preference: {str(e)}")


@app.post("/api/v2/survey/identity/reveal")
async def handle_identity_reveal(request: IdentityRevealRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Handle user request to reveal identity.

    Supports multiple reveal methods:
    - full: Complete identity reveal
    - partial_role: Reveal role/title only
    - partial_department: Reveal department only
    - gradual: Gradual reveal over time
    - consent_based: Consent-based reveal
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        identity_manager = SurveyIdentityManager(db)
        reveal_request = {
            "method": request.method,
            "target": request.target,
            "consent": request.consent,
            "conditions": request.conditions,
            "revoke_anonymity": request.revoke_anonymity,
        }
        result = identity_manager.handle_identity_reveal(
            user_id=current_user_email, reveal_request=reveal_request, survey_id=request.survey_id
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error handling identity reveal: {str(e)}")


@app.get("/api/v2/survey/identity/status/{user_email}")
async def get_identity_status(
    user_email: EmailStr,
    http_request: Request,
    survey_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Get current identity status for a user.

    Returns:
    - current_mode: Current identity mode
    - anonymous_mode: Whether user is in anonymous mode
    - privacy_level: Current privacy level
    - reveal_options: Available reveal options
    - retention_policy: Data retention policy
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        if user_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)
        identity_manager = SurveyIdentityManager(db)
        result = identity_manager.get_identity_status(user_id=user_email, survey_id=survey_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting identity status: {str(e)}")


@app.post("/api/v2/survey/identity/revoke-anonymity")
async def revoke_anonymity(
    http_request: Request,
    user_email: Optional[EmailStr] = None,
    survey_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Revoke anonymity and switch to conditional/identified mode.

    This allows users to transition from anonymous to identified/conditional mode.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        target_email = str(user_email) if user_email else current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)
        identity_manager = SurveyIdentityManager(db)
        result = identity_manager.process_revoke_anonymity(user_id=target_email, survey_id=survey_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error revoking anonymity: {str(e)}")


class ConditionalRevealRequest(BaseModel):
    """Request model for conditional reveal configuration"""

    user_email: Optional[EmailStr] = None  # derived from auth in production
    reveal_after_survey: Optional[Dict[str, Any]] = None
    reveal_to_specific_people: Optional[Dict[str, Any]] = None
    time_based_reveal: Optional[Dict[str, Any]] = None
    consent_based_reveal: Optional[Dict[str, Any]] = None
    notify_before_reveal: Optional[bool] = True
    notify_after_reveal: Optional[bool] = True
    notify_on_condition_met: Optional[bool] = True
    enable_reminders: Optional[bool] = False
    notify_days_before: Optional[int] = 1
    reminder_frequency: Optional[str] = "weekly"
    survey_id: Optional[int] = None


@app.post("/api/v2/survey/identity/conditional-reveal")
async def process_conditional_reveal(
    request: ConditionalRevealRequest, http_request: Request, db: Session = Depends(get_db)
):
    """
    Process conditional reveal preferences.

    Supports multiple conditional reveal scenarios:
    - reveal_after_survey: Reveal after survey completion + cooling period
    - reveal_to_specific_people: Reveal to specific recipients with partial reveal
    - time_based_reveal: Time-based reveal after specified days
    - consent_based_reveal: Consent-based reveal with explicit consent
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        engine = ConditionalAnonymityEngine(db)

        user_choice = {
            "reveal_after_survey": request.reveal_after_survey or {},
            "reveal_to_specific_people": request.reveal_to_specific_people or {},
            "time_based_reveal": request.time_based_reveal or {},
            "consent_based_reveal": request.consent_based_reveal or {},
            "notify_before_reveal": request.notify_before_reveal,
            "notify_after_reveal": request.notify_after_reveal,
            "notify_on_condition_met": request.notify_on_condition_met,
            "enable_reminders": request.enable_reminders,
            "notify_days_before": request.notify_days_before,
            "reminder_frequency": request.reminder_frequency,
        }

        result = engine.process_conditional_reveal(user_id=current_user_email, user_choice=user_choice, survey_id=request.survey_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing conditional reveal: {str(e)}")


@app.get("/api/v2/survey/identity/conditional-reveal/check-triggers/{user_email}")
async def check_trigger_conditions(
    user_email: EmailStr,
    http_request: Request,
    survey_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Check if any trigger conditions have been met for conditional reveals.

    Returns:
    - triggers_met: List of triggers that have been activated
    - actions_required: Actions that need to be taken
    - status: Current status (active, pending, no_config)
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        if user_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)
        engine = ConditionalAnonymityEngine(db)
        result = engine.check_trigger_conditions(user_id=user_email, survey_id=survey_id)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking trigger conditions: {str(e)}")


@app.post("/api/v2/survey/identity/conditional-reveal/execute/{user_email}")
async def execute_conditional_reveal(
    user_email: EmailStr,
    http_request: Request,
    trigger: str = Query(
        ..., description="Trigger type: survey_completed, cooling_period_passed, time_based, manual_request, consent_received"
    ),
    survey_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Execute conditional reveal based on trigger.

    This endpoint is called when a trigger condition has been met.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        if user_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)
        from backend.conditional_anonymity_engine import RevealTrigger

        try:
            trigger_enum = RevealTrigger(trigger)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid trigger: {trigger}")

        engine = ConditionalAnonymityEngine(db)
        result = engine.execute_conditional_reveal(user_id=user_email, trigger=trigger_enum, survey_id=survey_id)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error executing conditional reveal: {str(e)}")


# ============================================================================
# Hybrid Identity Survey System Endpoints
# ============================================================================


class HybridSessionRequest(BaseModel):
    """Request model for initializing hybrid identity session"""

    user_email: Optional[EmailStr] = None
    user_id: Optional[str] = None  # Alias for user_email
    preferred_mode: Optional[str] = Field(None, description="Identity mode: anonymous, conditional, partial, identified")
    identity_mode: Optional[str] = None  # Alias for preferred_mode
    survey_id: Optional[int] = None

    def get_user_email(self):
        """Get user email from either user_email or user_id"""
        return self.user_email or self.user_id

    def get_preferred_mode(self):
        """Get preferred mode from either preferred_mode or identity_mode"""
        return self.preferred_mode or self.identity_mode


class SurveyResponseRequest(BaseModel):
    """Request model for survey response submission"""

    session_token: str
    responses: Optional[Dict[str, Any]] = None
    question_id: Optional[int] = None
    response_text: Optional[str] = None
    response_value: Optional[Dict[str, Any]] = None
    survey_id: Optional[int] = None
    reveal_conditions: Optional[Dict[str, Any]] = None


class ModeSwitchRequest(BaseModel):
    """Request model for switching identity modes"""

    user_email: Optional[EmailStr] = None  # derived from auth in production
    new_mode: str
    reason: Optional[str] = None


@app.post("/api/v2/hybrid-identity/initialize-session")
async def initialize_hybrid_session(request: HybridSessionRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Initialize a new user session with chosen identity mode.

    Supports 4 identity modes:
    - anonymous: Complete anonymity
    - conditional: Conditional anonymity with reveal options
    - partial: Partially identified (role/department only)
    - identified: Fully identified
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        hybrid_system = HybridIdentitySurveySystem(db)
        result = hybrid_system.initialize_user_session(
            user_id=current_user_email, preferred_mode=request.preferred_mode, survey_id=request.survey_id
        )

        # Persist the session for cross-request reliability.
        session_token = result.get("session_token")
        if session_token:
            session_data = hybrid_system.sessions.get(session_token, {})
            expires_at = datetime.utcnow() + timedelta(hours=24)
            _store_hybrid_session(
                db,
                session_token=session_token,
                user_email=current_user_email,
                identity_mode=session_data.get("identity_mode") or result.get("mode") or request.preferred_mode,
                survey_id=session_data.get("survey_id"),
                permissions=session_data.get("permissions"),
                consent_granted=session_data.get("consent_granted"),
                expires_at=expires_at,
            )
            db.commit()

        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error initializing session: {str(e)}")


@app.post("/api/v2/hybrid-identity/create-survey-session")
async def create_survey_session(
    http_request: Request,
    user_email: Optional[EmailStr] = None,
    survey_type: str = Query(..., description="Survey type: comprehensive, climate, feedback, etc."),
    session_token: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Create a survey session based on user's identity mode.
    Returns mode-specific questions and privacy controls.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        target_email = str(user_email) if user_email else current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)

        hybrid_system = HybridIdentitySurveySystem(db)

        # Get user profile from session or create new
        if session_token:
            session_row = _get_hybrid_session(db, session_token)
            if not session_row:
                raise HTTPException(status_code=404, detail="Session not found")
            if session_row.expires_at and datetime.utcnow() > session_row.expires_at:
                raise HTTPException(status_code=404, detail="Session expired")
            if session_row.user_email.lower() != target_email.lower():
                _require_admin_access(http_request, db)

            session_row.last_activity = datetime.utcnow()
            db.commit()

            user_profile = {
                "user_id": target_email,
                "identity_mode": session_row.identity_mode,
                "anonymous_id": None,
            }
        else:
            # Create default profile
            user_profile = {"user_id": target_email, "identity_mode": "conditional", "anonymous_id": None}

        survey_session = hybrid_system.survey_engine.create_survey_session(user_profile=user_profile, survey_type=survey_type)
        return survey_session
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating survey session: {str(e)}")


@app.post("/api/v2/hybrid-identity/submit-response")
async def submit_survey_response(request: SurveyResponseRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Submit survey response with identity mode-specific processing.
    Applies anonymization, sentiment analysis, and theme extraction.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        hybrid_system = HybridIdentitySurveySystem(db)

        # Get session to determine identity mode
        session_row = _get_hybrid_session(db, request.session_token)
        if not session_row:
            raise HTTPException(status_code=404, detail="Session not found")

        if session_row.expires_at and datetime.utcnow() > session_row.expires_at:
            raise HTTPException(status_code=404, detail="Session expired")

        if session_row.user_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)

        session_row.last_activity = datetime.utcnow()
        db.commit()

        identity_mode = HybridIdentityMode(session_row.identity_mode or "conditional")

        # Handle both single response and batch responses
        if request.question_id:
            # Single response format
            response_data = {
                "session_id": request.session_token,
                "responses": {
                    str(request.question_id): {
                        "question_id": request.question_id,
                        "response_text": request.response_text,
                        "response_value": request.response_value,
                    }
                },
                "reveal_conditions": request.reveal_conditions,
            }
        else:
            # Batch responses format
            response_data = {
                "session_id": request.session_token,
                "responses": request.responses or {},
                "reveal_conditions": request.reveal_conditions,
            }

        processed_response = hybrid_system.survey_engine.process_response(
            response_data=response_data, identity_mode=identity_mode
        )
        return processed_response
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing response: {str(e)}")


@app.post("/api/v2/hybrid-identity/switch-mode")
async def switch_identity_mode(request: ModeSwitchRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Switch user's identity mode.
    Handles data migration and privacy changes.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        hybrid_system = HybridIdentitySurveySystem(db)

        try:
            new_mode = HybridIdentityMode(request.new_mode.lower())
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid identity mode: {request.new_mode}")

        result = hybrid_system.identity_manager.switch_identity_mode(
            user_id=current_user_email, new_mode=new_mode, reason=request.reason or ""
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error switching mode: {str(e)}")


@app.post("/api/v2/hybrid-identity/process-reveal-request")
async def process_reveal_request(
    user_email: EmailStr,
    http_request: Request,
    reveal_type: str = Query(..., description="Reveal type: full, partial, conditional"),
    conditions: Optional[Dict[str, Any]] = None,
    db: Session = Depends(get_db),
):
    """
    Process user request to reveal identity.
    Supports full, partial, and conditional reveals.
    """
    try:
        current_user_email = _require_authenticated_email(http_request)
        if user_email.lower() != current_user_email.lower():
            _require_admin_access(http_request, db)
        hybrid_system = HybridIdentitySurveySystem(db)
        result = hybrid_system.identity_manager.process_reveal_request(
            user_id=user_email, reveal_type=reveal_type, conditions=conditions or {}
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing reveal request: {str(e)}")


@app.get("/api/v2/hybrid-identity/analyze-survey-data")
async def analyze_survey_data(http_request: Request, survey_id: Optional[int] = Query(None), db: Session = Depends(get_db)):
    """
    Comprehensive analysis across all identity modes.
    Includes bias detection, trend analysis, and predictive insights.
    """
    try:
        _require_admin_access(http_request, db)
        hybrid_system = HybridIdentitySurveySystem(db)
        survey_data = hybrid_system.get_all_survey_data()
        identity_breakdown = hybrid_system.get_identity_breakdown()

        analysis_results = hybrid_system.analytics_engine.analyze_survey_data(
            survey_data=survey_data, identity_breakdown=identity_breakdown
        )
        return analysis_results
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing survey data: {str(e)}")


# ============================================================================
# Survey Management Endpoints
# ============================================================================


class SurveyCreateRequest(BaseModel):
    """Request model for creating a survey"""

    title: str
    description: Optional[str] = None
    survey_type: str = "comprehensive"
    start_date: Optional[date] = None
    end_date: Optional[date] = None


class SurveyUpdateRequest(BaseModel):
    """Request model for updating a survey"""

    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None


@app.get("/api/v2/surveys")
async def get_all_surveys(
    status: Optional[str] = Query(None, description="Filter by status"),
    survey_type: Optional[str] = Query(None, description="Filter by survey type"),
    db: Session = Depends(get_db),
):
    """
    List all surveys.
    Returns surveys filtered by status and type if provided.
    """
    try:
        query = db.query(Survey)
        if status:
            query = query.filter(Survey.status == status)
        if survey_type:
            query = query.filter(Survey.survey_type == survey_type)
        surveys = query.order_by(Survey.created_at.desc()).all()
        return [
            {
                "id": s.id,
                "title": s.title,
                "description": s.description,
                "survey_type": s.survey_type,
                "status": s.status,
                "start_date": s.start_date.isoformat() if s.start_date else None,
                "end_date": s.end_date.isoformat() if s.end_date else None,
                "created_at": s.created_at.isoformat(),
            }
            for s in surveys
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching surveys: {str(e)}")


@app.get("/api/v2/surveys/{survey_id}")
async def get_survey_details(survey_id: int, db: Session = Depends(get_db)):
    """Get survey details including questions count"""
    try:
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        questions_count = db.query(SurveyQuestion).filter(SurveyQuestion.survey_id == survey_id).count()
        responses_count = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey_id).count()

        return {
            "id": survey.id,
            "title": survey.title,
            "description": survey.description,
            "survey_type": survey.survey_type,
            "status": survey.status,
            "start_date": survey.start_date.isoformat() if survey.start_date else None,
            "end_date": survey.end_date.isoformat() if survey.end_date else None,
            "questions_count": questions_count,
            "responses_count": responses_count,
            "created_at": survey.created_at.isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching survey: {str(e)}")


@app.post("/api/v2/surveys")
async def create_survey(request: SurveyCreateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Create a new survey"""
    try:
        current_user_email = _require_admin_access(http_request, db)
        survey = Survey(
            title=request.title,
            description=request.description,
            survey_type=request.survey_type,
            start_date=request.start_date,
            end_date=request.end_date,
            created_by=current_user_email,
            status="draft",
        )
        db.add(survey)
        db.commit()
        db.refresh(survey)
        return {"id": survey.id, "title": survey.title, "status": survey.status, "message": "Survey created successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating survey: {str(e)}")


@app.put("/api/v2/surveys/{survey_id}")
async def update_survey(survey_id: int, request: SurveyUpdateRequest, http_request: Request, db: Session = Depends(get_db)):
    """Update an existing survey"""
    try:
        _require_admin_access(http_request, db)
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        if request.title is not None:
            survey.title = request.title
        if request.description is not None:
            survey.description = request.description
        if request.status is not None:
            survey.status = request.status
        if request.start_date is not None:
            survey.start_date = request.start_date
        if request.end_date is not None:
            survey.end_date = request.end_date

        db.commit()
        return {"message": "Survey updated successfully", "id": survey.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating survey: {str(e)}")


@app.get("/api/v2/surveys/{survey_id}/questions")
async def get_survey_questions(survey_id: int, db: Session = Depends(get_db)):
    """Get all questions for a survey"""
    try:
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        questions = (
            db.query(SurveyQuestion).filter(SurveyQuestion.survey_id == survey_id).order_by(SurveyQuestion.order_index).all()
        )

        return [
            {
                "id": q.id,
                "question_text": q.question_text,
                "question_type": q.question_type,
                "category": q.category,
                "section": q.section,
                "order_index": q.order_index,
                "required": q.required,
                "identity_modes": q.identity_modes,
                "sensitivity_level": q.sensitivity_level,
                "options": q.options,
            }
            for q in questions
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching questions: {str(e)}")


@app.get("/api/v2/surveys/{survey_id}/responses")
async def get_survey_responses(
    survey_id: int,
    http_request: Request,
    user_email: Optional[EmailStr] = Query(None, description="(Admin only) filter responses for a specific user"),
    db: Session = Depends(get_db),
):
    """
    Get survey responses (with permissions).
    Users can only see their own responses unless they're admins.
    """
    try:
        _require_admin_access(http_request, db)
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        query = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey_id)
        if user_email:
            query = query.filter(SurveyResponse.respondent_email == str(user_email))

        responses = query.all()

        return [
            {
                "id": r.id,
                "question_id": r.question_id,
                "respondent_email": r.respondent_email,
                "identity_mode": r.identity_mode,
                "response_text": r.response_text,
                "response_value": r.response_value,
                "submitted_at": r.submitted_at.isoformat(),
            }
            for r in responses
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching responses: {str(e)}")


class SurveyResponseSubmitRequest(BaseModel):
    """Request model for submitting survey responses"""

    survey_id: int
    question_id: int
    respondent_email: Optional[EmailStr] = None
    anonymous_id: Optional[str] = None
    session_id: Optional[str] = None
    identity_mode: str
    response_text: Optional[str] = None
    response_value: Optional[Dict[str, Any]] = None


@app.post("/api/v2/surveys/responses")
async def submit_survey_response_direct(
    request: SurveyResponseSubmitRequest, http_request: Request, db: Session = Depends(get_db)
):
    """Submit a survey response directly"""
    try:
        current_user_email = _require_authenticated_email(http_request)
        # Verify survey exists
        survey = db.query(Survey).filter(Survey.id == request.survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        # Verify question exists
        question = (
            db.query(SurveyQuestion)
            .filter(SurveyQuestion.id == request.question_id, SurveyQuestion.survey_id == request.survey_id)
            .first()
        )
        if not question:
            raise HTTPException(status_code=404, detail="Question not found")

        identity_mode = (request.identity_mode or "").lower()
        respondent_email = current_user_email if identity_mode == "identified" else None

        # Create response
        response = SurveyResponse(
            survey_id=request.survey_id,
            question_id=request.question_id,
            respondent_email=respondent_email,
            anonymous_id=request.anonymous_id,
            session_id=request.session_id,
            identity_mode=identity_mode,
            response_text=request.response_text,
            response_value=request.response_value,
        )
        db.add(response)
        db.commit()
        db.refresh(response)

        return {"message": "Response submitted successfully", "id": response.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error submitting response: {str(e)}")


@app.get("/api/v2/surveys/{survey_id}/analytics")
async def get_survey_analytics(survey_id: int, http_request: Request, db: Session = Depends(get_db)):
    """Get survey analytics and statistics"""
    try:
        _require_admin_access(http_request, db)
        survey = db.query(Survey).filter(Survey.id == survey_id).first()
        if not survey:
            raise HTTPException(status_code=404, detail="Survey not found")

        total_responses = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey_id).count()
        questions_count = db.query(SurveyQuestion).filter(SurveyQuestion.survey_id == survey_id).count()

        # Count by identity mode
        identity_mode_counts = (
            db.query(SurveyResponse.identity_mode, func.count(SurveyResponse.id))
            .filter(SurveyResponse.survey_id == survey_id)
            .group_by(SurveyResponse.identity_mode)
            .all()
        )

        return {
            "survey_id": survey_id,
            "total_responses": total_responses,
            "questions_count": questions_count,
            "response_rate": (total_responses / questions_count * 100) if questions_count > 0 else 0,
            "identity_mode_breakdown": {mode: count for mode, count in identity_mode_counts},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching analytics: {str(e)}")


# Notifications Endpoints (In-App)
# ============================================================================


@app.get("/api/v2/notifications/unread-count")
async def get_unread_notification_count(
    request: Request,
    user_email: Optional[EmailStr] = Query(None, description="(Admin only) fetch count for a specific user"),
    db: Session = Depends(get_db),
):
    """Get unread notification count for user"""
    try:
        current_user_email = _require_authenticated_email(request)
        target_email = str(user_email) if user_email else current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)

        count = (
            db.query(Notification)
            .filter(Notification.recipient_email == target_email, Notification.read == False)
            .count()
        )
        return {"unread_count": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching unread count: {str(e)}")


@app.post("/api/v2/notifications/mark-read")
async def mark_notifications_read(
    notification_ids: List[int],
    request: Request,
    user_email: Optional[EmailStr] = Query(None, description="(Admin only) mark notifications for a specific user"),
    db: Session = Depends(get_db),
):
    """Mark multiple notifications as read"""
    try:
        current_user_email = _require_authenticated_email(request)
        target_email = str(user_email) if user_email else current_user_email
        if target_email.lower() != current_user_email.lower():
            _require_admin_access(request, db)

        updated = (
            db.query(Notification)
            .filter(Notification.id.in_(notification_ids), Notification.recipient_email == target_email)
            .update({"read": True, "read_at": datetime.utcnow()}, synchronize_session=False)
        )
        db.commit()
        return {"message": f"Marked {updated} notifications as read"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error marking notifications as read: {str(e)}")


# Objections Endpoints
# ============================================================================


@app.get("/api/v2/objections/{objection_id}")
async def get_objection(objection_id: int, db: Session = Depends(get_db)):
    """Get a single objection by ID"""
    try:
        objection = db.query(Objection).filter(Objection.id == objection_id).first()
        if not objection:
            raise HTTPException(status_code=404, detail="Objection not found")

        return {
            "id": objection.id,
            "submitted_by": objection.submitted_by,
            "objection_type": objection.objection_type,
            "related_entity_type": objection.related_entity_type,
            "related_entity_id": objection.related_entity_id,
            "title": objection.title,
            "description": objection.description,
            "status": objection.status,
            "resolution_notes": objection.resolution_notes,
            "resolved_by": objection.resolved_by,
            "resolved_at": objection.resolved_at.isoformat() if objection.resolved_at else None,
            "created_at": objection.created_at.isoformat(),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching objection: {str(e)}")


# Reports Endpoints
# ============================================================================


@app.get("/api/v2/reports/bias/{cycle_id}")
async def get_bias_report(cycle_id: int, db: Session = Depends(get_db)):
    """Get bias report for a cycle"""
    try:
        bias_detection = Complete360BiasDetection(db)
        report = bias_detection.generate_comprehensive_report(cycle_id)
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating bias report: {str(e)}")


@app.get("/api/v2/reports/participation/{cycle_id}")
async def get_participation_report(cycle_id: int, db: Session = Depends(get_db)):
    """Get participation report for a cycle"""
    try:
        analytics = ParticipationAnalytics(db)
        report = analytics.analyze_participation(cycle_id)
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating participation report: {str(e)}")


# Analytics Endpoints
# ============================================================================


@app.get("/api/v2/analytics/bias/{cycle_id}")
async def get_bias_analytics(cycle_id: int, db: Session = Depends(get_db)):
    """Get bias analytics for a cycle"""
    try:
        bias_analytics = BiasAnalytics(db)
        analytics = bias_analytics.generate_analytics_report(cycle_id)
        return analytics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating bias analytics: {str(e)}")


@app.get("/api/v2/analytics/eom/{cycle_id}")
async def get_eom_analytics(cycle_id: int, db: Session = Depends(get_db)):
    """Get EOM analytics for a cycle"""
    try:
        eom_cycle = db.query(EOMCycle).filter(EOMCycle.cycle_id == cycle_id).first()
        if not eom_cycle:
            raise HTTPException(status_code=404, detail="EOM cycle not found")

        total_nominations = db.query(EOMNominee).filter(EOMNominee.eom_cycle_id == eom_cycle.id).count()
        total_voters = db.query(EOMVoter).filter(EOMVoter.eom_cycle_id == eom_cycle.id).count()

        # Count by category
        category_counts = (
            db.query(EOMNominee.category, func.count(EOMNominee.id))
            .filter(EOMNominee.eom_cycle_id == eom_cycle.id)
            .group_by(EOMNominee.category)
            .all()
        )

        return {
            "cycle_id": cycle_id,
            "eom_cycle_id": eom_cycle.id,
            "total_nominations": total_nominations,
            "total_voters": total_voters,
            "category_breakdown": {str(cat): count for cat, count in category_counts},
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating EOM analytics: {str(e)}")


@app.get("/api/v2/analytics/mre/{cycle_id}")
async def get_mre_analytics(cycle_id: int, db: Session = Depends(get_db)):
    """Get MRE analytics for a cycle"""
    try:
        total_assignments = db.query(Assignment).filter(Assignment.cycle_id == cycle_id).count()
        total_evaluations = db.query(Evaluation).join(Assignment).filter(Assignment.cycle_id == cycle_id).count()

        completion_rate = (total_evaluations / total_assignments * 100) if total_assignments > 0 else 0

        return {
            "cycle_id": cycle_id,
            "total_assignments": total_assignments,
            "total_evaluations": total_evaluations,
            "completion_rate": round(completion_rate, 2),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating MRE analytics: {str(e)}")


# Survey Templates Endpoints
# ============================================================================


@app.get("/api/v2/survey-templates/comprehensive")
async def get_comprehensive_survey_template(
    identity_mode: str = Query(..., description="Identity mode: anonymous, conditional, partial, identified"),
    db: Session = Depends(get_db),
):
    """
    Get comprehensive school climate survey template.
    Returns mode-specific questions and sections.
    """
    try:
        templates = EternitySchoolSurveyTemplates(db)
        survey_template = templates.get_comprehensive_school_survey(identity_mode=identity_mode, survey_type="comprehensive")
        return survey_template
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting survey template: {str(e)}")


@app.get("/api/v2/survey-templates/section/{category}")
async def get_survey_section(
    category: str, identity_mode: str = Query(..., description="Identity mode"), db: Session = Depends(get_db)
):
    """
    Get specific survey section by category.
    Categories: physical_environment, workplace_culture, management, etc.
    """
    try:
        templates = EternitySchoolSurveyTemplates(db)
        section = templates.get_template_by_category(category, identity_mode)
        return section
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting survey section: {str(e)}")


# ============================================================================
# Admin Dashboard Endpoints
# ============================================================================


@app.get("/api/v2/admin/dashboard")
async def get_admin_dashboard(
    request: Request,
    admin_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """
    Get comprehensive admin dashboard data.
    Includes overview cards, real-time metrics, analytics, and action items.
    """
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(request)
        rbac = RBACSystem(db)
        role = rbac.get_user_role(current_user_email)
        if not (rbac.is_super_admin(current_user_email) or role in ("ceo", "pnc")):
            raise HTTPException(status_code=403, detail="Admin access required")

        # Backwards compatibility: if admin_id is provided, restrict non-super-admin users to themselves
        if admin_id and admin_id.lower() != current_user_email.lower() and not rbac.is_super_admin(current_user_email):
            raise HTTPException(status_code=403, detail="You can only view your own admin dashboard")

        dashboard = EternitySchoolAdminDashboard(db)
        dashboard_data = dashboard.get_main_dashboard(current_user_email)
        return dashboard_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting admin dashboard: {str(e)}")


@app.get("/api/v2/admin/dashboard/overview-cards")
async def get_overview_cards(request: Request, db: Session = Depends(get_db)):
    """Get overview metric cards for dashboard"""
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(request)
        rbac = RBACSystem(db)
        role = rbac.get_user_role(current_user_email)
        if not (rbac.is_super_admin(current_user_email) or role in ("ceo", "pnc")):
            raise HTTPException(status_code=403, detail="Admin access required")

        dashboard = EternitySchoolAdminDashboard(db)
        cards = dashboard.get_overview_cards()
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting overview cards: {str(e)}")


@app.get("/api/v2/admin/dashboard/real-time-metrics")
async def get_real_time_metrics(request: Request, db: Session = Depends(get_db)):
    """Get real-time system metrics"""
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(request)
        rbac = RBACSystem(db)
        role = rbac.get_user_role(current_user_email)
        if not (rbac.is_super_admin(current_user_email) or role in ("ceo", "pnc")):
            raise HTTPException(status_code=403, detail="Admin access required")

        dashboard = EternitySchoolAdminDashboard(db)
        metrics = dashboard.get_real_time_metrics()
        return metrics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting real-time metrics: {str(e)}")


@app.get("/api/v2/admin/dashboard/identity-analytics")
async def get_identity_analytics(request: Request, db: Session = Depends(get_db)):
    """Get detailed identity mode analytics"""
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(request)
        rbac = RBACSystem(db)
        role = rbac.get_user_role(current_user_email)
        if not (rbac.is_super_admin(current_user_email) or role in ("ceo", "pnc")):
            raise HTTPException(status_code=403, detail="Admin access required")

        dashboard = EternitySchoolAdminDashboard(db)
        analytics = dashboard.get_identity_mode_analytics()
        return analytics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting identity analytics: {str(e)}")


# ============================================================================
# Admin Settings Endpoints (CEO / Super Admin)
# ============================================================================


class SystemSettingsPayload(BaseModel):
    email_notifications: Optional[bool] = None
    auto_activate_cycles: Optional[bool] = None
    require_approval: Optional[bool] = None
    default_rotation_period: Optional[str] = None  # term, quarter, month, year
    max_nominations_per_person: Optional[int] = None
    evaluation_deadline_days: Optional[int] = None

    @field_validator("default_rotation_period")
    @classmethod
    def _validate_rotation_period(cls, v: Optional[str]):
        if v is None:
            return v
        allowed = {"term", "quarter", "month", "year"}
        value = str(v).strip().lower()
        if value not in allowed:
            raise ValueError(f"default_rotation_period must be one of: {', '.join(sorted(allowed))}")
        return value

    @field_validator("max_nominations_per_person")
    @classmethod
    def _validate_max_nominations(cls, v: Optional[int]):
        if v is None:
            return v
        if v < 1 or v > 50:
            raise ValueError("max_nominations_per_person must be between 1 and 50")
        return v

    @field_validator("evaluation_deadline_days")
    @classmethod
    def _validate_deadline_days(cls, v: Optional[int]):
        if v is None:
            return v
        if v < 1 or v > 365:
            raise ValueError("evaluation_deadline_days must be between 1 and 365")
        return v


def _require_super_admin(request: Request, db: Session) -> str:
    from backend.rbac_system import RBACSystem

    current_user_email = _require_authenticated_email(request)
    rbac = RBACSystem(db)
    if not rbac.is_super_admin(current_user_email):
        raise HTTPException(status_code=403, detail="Super admin access required")
    return current_user_email


@app.get("/api/v2/admin/settings")
async def get_system_settings(request: Request, db: Session = Depends(get_db)):
    """Get global system settings (CEO / super admin only)."""
    try:
        _require_super_admin(request, db)

        settings = db.query(SystemSetting).filter(SystemSetting.id == 1).first()
        if not settings:
            raise HTTPException(status_code=404, detail="Settings not found")

        return {
            "email_notifications": settings.email_notifications,
            "auto_activate_cycles": settings.auto_activate_cycles,
            "require_approval": settings.require_approval,
            "default_rotation_period": settings.default_rotation_period,
            "max_nominations_per_person": settings.max_nominations_per_person,
            "evaluation_deadline_days": settings.evaluation_deadline_days,
            "updated_by": settings.updated_by,
            "updated_at": settings.updated_at.isoformat() if settings.updated_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting settings: {str(e)}")


@app.put("/api/v2/admin/settings")
async def update_system_settings(payload: SystemSettingsPayload, request: Request, db: Session = Depends(get_db)):
    """Update global system settings (CEO / super admin only)."""
    try:
        current_user_email = _require_super_admin(request, db)

        settings = db.query(SystemSetting).filter(SystemSetting.id == 1).first()
        if not settings:
            raise HTTPException(status_code=404, detail="Settings not found")

        for key, value in payload.model_dump(exclude_unset=True).items():
            if value is not None:
                setattr(settings, key, value)

        settings.updated_by = current_user_email
        db.commit()
        db.refresh(settings)

        return {
            "email_notifications": settings.email_notifications,
            "auto_activate_cycles": settings.auto_activate_cycles,
            "require_approval": settings.require_approval,
            "default_rotation_period": settings.default_rotation_period,
            "max_nominations_per_person": settings.max_nominations_per_person,
            "evaluation_deadline_days": settings.evaluation_deadline_days,
            "updated_by": settings.updated_by,
            "updated_at": settings.updated_at.isoformat() if settings.updated_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating settings: {str(e)}")


@app.post("/api/v2/admin/settings")
async def create_system_settings(payload: SystemSettingsPayload, request: Request, db: Session = Depends(get_db)):
    """Create global system settings if missing (CEO / super admin only)."""
    try:
        current_user_email = _require_super_admin(request, db)

        existing = db.query(SystemSetting).filter(SystemSetting.id == 1).first()
        if existing:
            # Idempotent: treat as update
            for key, value in payload.model_dump(exclude_unset=True).items():
                if value is not None:
                    setattr(existing, key, value)
            existing.updated_by = current_user_email
            db.commit()
            db.refresh(existing)
            return {
                "email_notifications": existing.email_notifications,
                "auto_activate_cycles": existing.auto_activate_cycles,
                "require_approval": existing.require_approval,
                "default_rotation_period": existing.default_rotation_period,
                "max_nominations_per_person": existing.max_nominations_per_person,
                "evaluation_deadline_days": existing.evaluation_deadline_days,
                "updated_by": existing.updated_by,
                "updated_at": existing.updated_at.isoformat() if existing.updated_at else None,
            }

        settings = SystemSetting(id=1)
        for key, value in payload.model_dump(exclude_unset=True).items():
            if value is not None:
                setattr(settings, key, value)
        settings.updated_by = current_user_email

        db.add(settings)
        db.commit()
        db.refresh(settings)

        return {
            "email_notifications": settings.email_notifications,
            "auto_activate_cycles": settings.auto_activate_cycles,
            "require_approval": settings.require_approval,
            "default_rotation_period": settings.default_rotation_period,
            "max_nominations_per_person": settings.max_nominations_per_person,
            "evaluation_deadline_days": settings.evaluation_deadline_days,
            "updated_by": settings.updated_by,
            "updated_at": settings.updated_at.isoformat() if settings.updated_at else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error saving settings: {str(e)}")


# ============================================================================
# Integration Hub Endpoints
# ============================================================================


class HRIntegrationConfig(BaseModel):
    """HR system integration configuration"""

    hr_system_url: str
    api_key: Optional[str] = None
    real_time_sync: Optional[bool] = False
    webhook_url: Optional[str] = None
    ip_whitelist: Optional[List[str]] = None


@app.post("/api/v2/integration/hr/setup")
async def setup_hr_integration(config: HRIntegrationConfig, http_request: Request, db: Session = Depends(get_db)):
    """
    Set up integration with HR system.
    Configures two-way sync, evaluation bridge, and security protocols.
    """
    try:
        _require_super_admin(http_request, db)
        integration_hub = EternitySchoolIntegrationHub(db)
        hr_config = {
            "hr_system_url": config.hr_system_url,
            "api_key": config.api_key,
            "real_time_sync": config.real_time_sync,
            "webhook_url": config.webhook_url,
            "ip_whitelist": config.ip_whitelist or [],
        }
        result = integration_hub.integrate_with_hr_system(hr_config)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting up HR integration: {str(e)}")


@app.get("/api/v2/integration/evaluation-bridge")
async def get_evaluation_bridge(http_request: Request, db: Session = Depends(get_db)):
    """
    Get evaluation data bridge configuration.
    Shows how survey feedback integrates with evaluation system.
    """
    try:
        _require_admin_access(http_request, db)
        integration_hub = EternitySchoolIntegrationHub(db)
        bridge = integration_hub.create_evaluation_data_bridge()
        return bridge
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting evaluation bridge: {str(e)}")


@app.post("/api/v2/integration/sync/staff")
async def sync_staff_data(staff_data: List[Dict[str, Any]], http_request: Request, db: Session = Depends(get_db)):
    """
    Sync staff data from HR system.
    Bidirectional sync of staff information.
    """
    try:
        _require_super_admin(http_request, db)
        integration_hub = EternitySchoolIntegrationHub(db)
        result = integration_hub.sync_staff_data(staff_data)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing staff data: {str(e)}")


@app.post("/api/v2/integration/sync/evaluation")
async def sync_evaluation_data(evaluation_data: Dict[str, Any], http_request: Request, db: Session = Depends(get_db)):
    """
    Sync evaluation data between systems.
    Handles survey feedback to evaluation integration.
    """
    try:
        _require_super_admin(http_request, db)
        integration_hub = EternitySchoolIntegrationHub(db)
        result = integration_hub.sync_evaluation_data(evaluation_data)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing evaluation data: {str(e)}")


# ============================================================================
# System Setup Endpoints
# ============================================================================


class SchoolConfigRequest(BaseModel):
    """School configuration request"""

    hr_sync_enabled: Optional[bool] = False
    hr_sync_frequency: Optional[str] = "daily"
    real_time_sync: Optional[bool] = False
    environment: Optional[str] = "production"


@app.post("/api/v2/system/setup")
async def setup_complete_system(config: SchoolConfigRequest, http_request: Request, db: Session = Depends(get_db)):
    """
    Set up the complete integrated system.
    Configures all components: templates, identity, dashboard, integrations.
    """
    try:
        _require_super_admin(http_request, db)
        system_setup = EternitySchoolSystemSetup(db)
        school_config = {
            "hr_sync_enabled": config.hr_sync_enabled,
            "hr_sync_frequency": config.hr_sync_frequency,
            "real_time_sync": config.real_time_sync,
            "environment": config.environment,
        }
        result = system_setup.setup_complete_system(school_config)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error setting up system: {str(e)}")


@app.get("/api/v2/system/go-live-checklist")
async def get_go_live_checklist(http_request: Request, db: Session = Depends(get_db)):
    """Get go-live checklist for system deployment"""
    try:
        _require_admin_access(http_request, db)
        system_setup = EternitySchoolSystemSetup(db)
        checklist = system_setup.generate_go_live_checklist()
        return checklist
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting checklist: {str(e)}")


# ============================================================================
# Audit Logs Endpoints
# ============================================================================


@app.get("/api/v2/audit-logs")
async def get_audit_logs(
    http_request: Request,
    action_type: Optional[str] = None,
    entity_type: Optional[str] = None,
    user_email: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
):
    """Get audit logs with filters"""
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)
        current_role = rbac.get_user_role(current_user_email)
        is_admin = rbac.is_super_admin(current_user_email) or current_role in ("ceo", "pnc")

        # Non-admins can only view their own logs.
        if not is_admin:
            user_email = current_user_email

        query = db.query(AuditLog)

        if entity_type:
            query = query.filter(AuditLog.entity_type == entity_type)

        if user_email:
            query = query.filter(AuditLog.user_email == user_email)

        if action_type:
            try:
                query = query.filter(AuditLog.action_type == DBActionType(action_type.lower()))
            except ValueError:
                raise HTTPException(status_code=422, detail="Invalid action_type")

        if date_from:
            try:
                start_dt = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
                query = query.filter(AuditLog.created_at >= start_dt)
            except ValueError:
                raise HTTPException(status_code=422, detail="Invalid date_from. Use ISO format.")

        if date_to:
            try:
                # If only a date is provided (no time), treat it as inclusive end-of-day.
                if "T" in date_to:
                    end_dt = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
                    query = query.filter(AuditLog.created_at <= end_dt)
                else:
                    end_date = date.fromisoformat(date_to)
                    query = query.filter(AuditLog.created_at < (datetime.combine(end_date, datetime.min.time()) + timedelta(days=1)))
            except ValueError:
                raise HTTPException(status_code=422, detail="Invalid date_to. Use ISO format.")

        logs = query.order_by(AuditLog.created_at.desc()).limit(limit).all()
        return [
            {
                "id": log.id,
                "action_type": log.action_type.value if hasattr(log.action_type, "value") else str(log.action_type),
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "user_email": log.user_email,
                "description": log.description,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# RBAC Permission Management Endpoints
# ============================================================================


class GrantPermissionRequest(BaseModel):
    """Request to grant a permission"""

    user_email: EmailStr
    permission_type: str
    expires_at: Optional[str] = None  # ISO format datetime, None = unlimited
    metadata: Optional[Dict[str, Any]] = None


class RevokePermissionRequest(BaseModel):
    """Request to revoke a permission"""

    user_email: EmailStr
    permission_type: str


@app.post("/api/v2/admin/permissions/grant")
async def grant_permission(
    request: GrantPermissionRequest,
    http_request: Request,
    db: Session = Depends(get_db),
):
    """
    Grant a permission to a user.
    Only super admin (bootstrap CEO) or users with GRANT_PERMISSIONS can use this.
    """
    try:
        from backend.rbac_system import PermissionType, RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)

        # Parse permission type
        try:
            permission = PermissionType(request.permission_type)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid permission type: {request.permission_type}")

        # Parse expiration date if provided
        expires_at = None
        if request.expires_at:
            try:
                expires_at = datetime.fromisoformat(request.expires_at.replace("Z", "+00:00"))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Invalid date format: {request.expires_at}. Use ISO format.")

        # Grant permission
        user_permission = rbac.grant_permission(
            user_email=request.user_email,
            permission=permission,
            granted_by=current_user_email,
            expires_at=expires_at,
            metadata=request.metadata,
        )

        return {
            "success": True,
            "permission": {
                "id": user_permission.id,
                "user_email": user_permission.user_email,
                "permission_type": user_permission.permission_type.value,
                "granted_by": user_permission.granted_by,
                "granted_at": user_permission.granted_at.isoformat(),
                "expires_at": user_permission.expires_at.isoformat() if user_permission.expires_at else None,
                "unlimited": user_permission.expires_at is None,
            },
            "message": f"Permission {permission.value} granted to {request.user_email}",
        }
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/admin/permissions/revoke")
async def revoke_permission(
    request: RevokePermissionRequest,
    http_request: Request,
    db: Session = Depends(get_db),
):
    """
    Revoke a permission from a user.
    Only super admin (bootstrap CEO) or users with REVOKE_PERMISSIONS can use this.
    """
    try:
        from backend.rbac_system import PermissionType, RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)

        # Parse permission type
        try:
            permission = PermissionType(request.permission_type)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid permission type: {request.permission_type}")

        # Revoke permission
        revoked = rbac.revoke_permission(user_email=request.user_email, permission=permission, revoked_by=current_user_email)

        if not revoked:
            raise HTTPException(
                status_code=404, detail=f"Active permission {permission.value} not found for {request.user_email}"
            )

        return {"success": True, "message": f"Permission {permission.value} revoked from {request.user_email}"}
    except PermissionError as e:
        raise HTTPException(status_code=403, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/admin/permissions/{user_email}")
async def get_user_permissions(
    user_email: str,
    http_request: Request,
    db: Session = Depends(get_db),
):
    """
    Get all permissions for a user.
    Users can view their own permissions, super admin can view anyone's.
    """
    try:
        from backend.rbac_system import RBACSystem

        current_user_email = _require_authenticated_email(http_request)
        rbac = RBACSystem(db)

        # Check if user can view permissions
        if user_email.lower() != current_user_email.lower() and not rbac.is_super_admin(current_user_email):
            raise HTTPException(status_code=403, detail="You can only view your own permissions")

        permissions = rbac.get_user_permissions(user_email)

        return {
            "user_email": user_email,
            "is_super_admin": rbac.is_super_admin(user_email),
            "role": rbac.get_user_role(user_email),
            "permissions": permissions,
            "total_permissions": len(permissions),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Evaluator Assignment Management Endpoints
# ============================================================================


class CreateAssignmentsRequest(BaseModel):
    """Request to create evaluator assignments for a staff member"""

    target_email: EmailStr
    cycle_id: int
    evaluator_overrides: Optional[Dict[str, str]] = None  # rater_context -> evaluator_email


class UpdateAssignmentsRequest(BaseModel):
    """Request to update evaluator assignments"""

    target_email: EmailStr
    cycle_id: int
    assignments: List[Dict[str, Any]]  # List of assignment updates


@app.post("/api/v2/staff/{email}/assign-evaluators")
async def create_evaluator_assignments(
    email: str, request: CreateAssignmentsRequest, http_request: Request, db: Session = Depends(get_db)
):
    """
    Create evaluator assignments for a staff member.
    Automatically assigns all required evaluators based on staff type (academic/admin).
    """
    try:
        _require_admin_access(http_request, db)
        from backend.evaluator_assignment_manager import EvaluatorAssignmentManager

        manager = EvaluatorAssignmentManager(db)

        assignments = manager.create_assignments_for_staff(
            target_email=request.target_email, cycle_id=request.cycle_id, evaluator_overrides=request.evaluator_overrides
        )

        return {
            "success": True,
            "target_email": request.target_email,
            "cycle_id": request.cycle_id,
            "assignments_created": len(assignments),
            "assignments": [
                {
                    "id": a.id,
                    "rater_email": a.rater_email,
                    "rater_context": a.rater_context,
                    "weight": a.weight,
                    "target_group": a.target_group,
                }
                for a in assignments
            ],
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/v2/staff/{email}/evaluators")
async def update_evaluator_assignments(
    email: str, request: UpdateAssignmentsRequest, http_request: Request, db: Session = Depends(get_db)
):
    """
    Update evaluator assignments for a staff member.
    Used when roles change or evaluators need manual adjustment.
    """
    try:
        _require_admin_access(http_request, db)
        from backend.evaluator_assignment_manager import EvaluatorAssignmentManager

        manager = EvaluatorAssignmentManager(db)

        assignments = manager.update_assignments_for_staff(
            target_email=request.target_email, cycle_id=request.cycle_id, updated_assignments=request.assignments
        )

        return {
            "success": True,
            "target_email": request.target_email,
            "cycle_id": request.cycle_id,
            "assignments_updated": len(assignments),
            "assignments": [
                {
                    "id": a.id,
                    "rater_email": a.rater_email,
                    "rater_context": a.rater_context,
                    "weight": a.weight,
                    "target_group": a.target_group,
                }
                for a in assignments
            ],
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/staff/{email}/evaluators")
async def get_staff_evaluators(
    email: str, http_request: Request, cycle_id: Optional[int] = Query(None), db: Session = Depends(get_db)
):
    """
    Get all evaluator assignments for a staff member.
    Shows who evaluates this person.
    """
    try:
        _require_admin_access(http_request, db)
        from backend.evaluator_assignment_manager import EvaluatorAssignmentManager

        manager = EvaluatorAssignmentManager(db)

        assignments = manager.get_assignments_for_staff(target_email=email, cycle_id=cycle_id)

        return {
            "target_email": email,
            "cycle_id": cycle_id,
            "assignments": [
                {
                    "id": a.id,
                    "rater_email": a.rater_email,
                    "rater_name": a.rater.full_name if a.rater else a.rater_email,
                    "rater_role": a.rater_role,
                    "rater_context": a.rater_context,
                    "weight": a.weight,
                    "target_group": a.target_group,
                    "cycle_id": a.cycle_id,
                }
                for a in assignments
            ],
            "total_evaluators": len(assignments),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/evaluation-matrix/{cycle_id}")
async def get_evaluation_matrix(cycle_id: int, http_request: Request, db: Session = Depends(get_db)):
    """
    Get complete evaluation matrix showing who evaluates whom.
    Shows all evaluation relationships in the system.
    """
    try:
        _require_admin_access(http_request, db)
        from backend.evaluator_assignment_manager import EvaluatorAssignmentManager

        manager = EvaluatorAssignmentManager(db)

        matrix = manager.get_evaluation_matrix(cycle_id)

        return matrix
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/staff/{email}/evaluation-status")
async def get_staff_evaluation_status(
    email: str, http_request: Request, cycle_id: Optional[int] = Query(None), db: Session = Depends(get_db)
):
    """
    Get evaluation status for a staff member.
    Shows who they evaluate and who evaluates them.
    """
    try:
        _require_admin_access(http_request, db)
        from backend.evaluator_assignment_manager import EvaluatorAssignmentManager

        manager = EvaluatorAssignmentManager(db)

        # Get assignments where this person is the target (who evaluates them)
        evaluated_by = manager.get_assignments_for_staff(email, cycle_id)

        # Get assignments where this person is the rater (who they evaluate)
        evaluating = db.query(Assignment).filter(Assignment.rater_email == email)
        if cycle_id:
            evaluating = evaluating.filter(Assignment.cycle_id == cycle_id)
        evaluating = evaluating.all()

        person = db.query(Person).filter(Person.email == email).first()
        staff_type = manager.get_staff_type(person) if person else "unknown"
        required_evaluators = manager.get_required_evaluators(staff_type) if person else []

        return {
            "staff_email": email,
            "staff_name": person.full_name if person else email,
            "staff_type": staff_type,
            "cycle_id": cycle_id,
            "evaluated_by": [
                {
                    "assignment_id": a.id,
                    "rater_email": a.rater_email,
                    "rater_name": a.rater.full_name if a.rater else a.rater_email,
                    "rater_position": a.rater.role_title if a.rater else None,
                    "rater_department": a.rater.department if a.rater else None,
                    "rater_context": a.rater_context,
                    "weight": a.weight,
                    "status": "assigned",
                }
                for a in evaluated_by
            ],
            "evaluating": [
                {
                    "assignment_id": a.id,
                    "target_email": a.target_email,
                    "target_name": a.target.full_name if a.target else a.target_email,
                    "target_position": a.target.role_title if a.target else None,
                    "target_department": a.target.department if a.target else None,
                    "rater_context": a.rater_context,
                    "weight": a.weight,
                    "status": "assigned",
                }
                for a in evaluating
            ],
            "required_evaluators": required_evaluators,
            "summary": {
                "total_evaluators": len(evaluated_by),
                "total_evaluating": len(evaluating),
                "required_count": len([e for e in required_evaluators if e.get("required", False)]),
            },
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# Bulk Import Endpoints
# ============================================================================


@app.post("/api/v2/import/staff")
async def import_staff(http_request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import staff from Excel file"""
    try:
        _require_admin_access(http_request, db)
        import os
        import tempfile

        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        try:
            importer = BulkImporter(db)
            result = importer.import_staff_from_excel(tmp_path)
            return result
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/import/eom-voters")
async def import_eom_voters(
    http_request: Request,
    file: UploadFile = File(...),
    cycle_id: int = Query(...),
    month: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
):
    """Import EOM voters from Excel file"""
    try:
        _require_admin_access(http_request, db)
        import os
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        try:
            importer = BulkImporter(db)
            result = importer.import_eom_voters_from_excel(tmp_path, cycle_id, month, year)
            return result
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/import/eom-candidates")
async def import_eom_candidates(http_request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import EOM candidates from Excel file"""
    try:
        _require_admin_access(http_request, db)
        import os
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        try:
            importer = BulkImporter(db)
            result = importer.import_eom_candidates_from_excel(tmp_path)
            return result
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/v2/import/weight-matrix")
async def import_weight_matrix(
    http_request: Request, file: UploadFile = File(...), cycle_id: int = Query(...), db: Session = Depends(get_db)
):
    """Import weight matrix from Excel file"""
    try:
        _require_admin_access(http_request, db)
        import os
        import tempfile

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name

        try:
            importer = BulkImporter(db)
            result = importer.import_weight_matrix_from_excel(tmp_path, cycle_id)
            return result
        finally:
            os.unlink(tmp_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/analytics/participation/{cycle_id}")
async def analyze_participation(cycle_id: int, db: Session = Depends(get_db)):
    """
    Analyze participation rates, engagement trends, outliers, and predict future participation.

    Returns:
        - participation_rates: Participation rates by department and segment
        - engagement_trends: Trend analysis over time
        - outlier_detection: Departments/segments with low participation
        - prediction: Predictions for future cycles
    """
    try:
        # Get cycle
        cycle = db.query(Cycle).filter(Cycle.id == cycle_id).first()
        if not cycle:
            raise HTTPException(status_code=404, detail=f"Cycle {cycle_id} not found")

        # Analyze participation
        analytics = ParticipationAnalytics(db)
        result = analytics.analyze_participation(cycle)

        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing participation: {str(e)}")


# ============================================================================
# EOM Hall of Fame and Diversity Endpoints
# ============================================================================


@app.get("/api/v2/eom/hall-of-fame")
async def get_hall_of_fame(
    category: Optional[str] = Query(None, description="Filter by category"),
    year: Optional[int] = Query(None, description="Filter by year"),
    segment: Optional[str] = Query(None, description="Filter by segment"),
    db: Session = Depends(get_db),
):
    """Get EOM Hall of Fame (winners history)"""
    try:
        query = """
            SELECT * FROM eom_hall_of_fame
            WHERE 1=1
        """
        params = {}

        if category:
            query += " AND category = :category"
            params["category"] = category
        if year:
            query += " AND EXTRACT(YEAR FROM cycle_start) = :year"
            params["year"] = year
        if segment:
            query += " AND segment = :segment"
            params["segment"] = segment

        query += " ORDER BY cycle_start DESC, category"

        result = db.execute(query, params)
        winners = [dict(row) for row in result]

        return {"data": winners}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/eom/diversity-tracking")
async def get_diversity_tracking(
    cycle_id: Optional[int] = Query(None, description="Filter by cycle ID"), db: Session = Depends(get_db)
):
    """Get EOM diversity tracking data"""
    try:
        query = """
            SELECT * FROM eom_diversity_tracking
            WHERE 1=1
        """
        params = {}

        if cycle_id:
            query += " AND cycle_id = :cycle_id"
            params["cycle_id"] = cycle_id

        result = db.execute(query, params)
        data = [dict(row) for row in result]

        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class EOMFeedbackSubmitRequest(BaseModel):
    eom_cycle_id: int
    feedback_type: str
    feedback_text: str
    rating: Optional[int] = Field(None, ge=1, le=5)
    person_email: Optional[str] = None  # ignored; derived from auth in production


@app.post("/api/v2/eom/feedback")
async def submit_eom_feedback(request: EOMFeedbackSubmitRequest, http_request: Request, db: Session = Depends(get_db)):
    """Submit EOM feedback"""
    try:
        from backend.database import EOMFeedback

        current_user_email = _require_authenticated_email(http_request)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, request.eom_cycle_id)

        feedback = EOMFeedback(
            eom_cycle_id=resolved_eom_cycle_id,
            feedback_type=request.feedback_type,
            person_email=current_user_email,
            feedback_text=request.feedback_text,
            rating=request.rating,
        )

        db.add(feedback)
        db.commit()
        db.refresh(feedback)

        return {"success": True, "feedback_id": feedback.id, "message": "Feedback submitted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v2/eom/cycles/{cycle_id}/window-status")
async def get_nomination_window_status(cycle_id: int, request: Request, db: Session = Depends(get_db)):
    """Get nomination window status for an EOM cycle"""
    try:
        from backend.eom_validation import EOMNominationValidator

        _require_authenticated_email(request)
        resolved_eom_cycle_id = _resolve_eom_cycle_id(db, cycle_id)
        eom_cycle = db.query(EOMCycle).filter(EOMCycle.id == resolved_eom_cycle_id).first()
        if not eom_cycle:
            raise HTTPException(status_code=404, detail="EOM cycle not found")

        validator = EOMNominationValidator(db)
        window_check = validator._check_nomination_window(eom_cycle)

        return window_check
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/")
async def root():
    """Root endpoint with API information and quick reference"""
    return {
        "message": "Eternity School Evaluation System API",
        "version": "2.0.0",
        "framework": "FastAPI",
        "documentation": {
            "swagger_ui": "/docs",
            "redoc": "/redoc",
            "app_index": "See docs/APP_INDEX.md for complete application index",
        },
        "quick_links": {
            "health": "/api/v2/health",
            "current_cycle": "/api/v2/cycles/current",
            "eom_nominations": "/api/v2/eom/nominations",
            "eom_hall_of_fame": "/api/v2/eom/hall-of-fame",
            "eom_diversity": "/api/v2/eom/diversity-tracking",
            "mre_evaluations": "/api/v2/mre/evaluations",
            "bias_reports": "/api/v2/bias/reports",
            "bias_analytics": "/api/v2/bias/analytics",
            "ceo_reports": "/api/v2/reports/ceo",
            "participation_analytics": "/api/v2/analytics/participation/{cycle_id}",
            "smart_notifications": "/api/v2/notifications/user-behavior/{email}",
            "objections": "/api/v2/objections",
            "audit_logs": "/api/v2/audit-logs",
        },
        "main_endpoints": {
            "cycles": "/api/v2/cycles",
            "eom": "/api/v2/eom",
            "mre": "/api/v2/mre",
            "bias": "/api/v2/bias",
            "scoring": "/api/v2/scoring",
            "reports": "/api/v2/reports",
            "notifications": "/api/v2/notifications",
            "analytics": "/api/v2/analytics",
            "import": "/api/v2/import",
        },
    }
